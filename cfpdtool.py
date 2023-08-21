#!/usr/bin/python 

# standard modules
import calendar
import datetime
import numpy
import math
import operator
import os
import pygame
import random
import sys
import wx
import xlrd
import openpyxl


# local modules
import blockanalysis
import conversion
import flowstatexports
import fsconfig
import mystat
import visual
import screenshot
import supportfunctions

class MyFrame(wx.Frame):
    def __init__(self):

         
        # version string
        self.version="Open Source release"
        self.appname="CFPD tool "+self.version

        self.cwd=os.getcwd()
        # application screen size
        self.appsize='large'

        # platform
        self.platform=os.name
        if self.platform=='posix':
            self.pathseparator='/'
        else:
            self.pathseparator='\\'

        # create a frame, no parent, default to wxID_ANY
        margin=4
        self.xsize=1024
        self.ysize=720#+64
        self.fppanelysize=109 # 125
        # fonts
        self.textfont = wx.Font(8, wx.SWISS, wx.NORMAL, wx.NORMAL, False)
        self.boldtextfont = wx.Font(8, wx.SWISS, wx.NORMAL, wx.BOLD, False)

        self.xframesize=self.xsize-2*margin -16
        self.xframesize2=0.75*self.xsize-2*margin -8
        self.xframesize3=0.25*self.xsize-2*margin -8

        # data arrays
        self.datx=[40182,40189]
        self.daty=[0,0]
        self.datt=[0,1]
        self.datx2=[40182,40189]
        self.daty2=[0,0]
        self.datt2=[0,1]
        self.minxzoom=min(self.datx)
        self.maxxzoom=max(self.datx)
        self.minxzoom2=min(self.datx2)
        self.maxxzoom2=max(self.datx2)

        self.data_Q=[]
        self.data_sigma=[]
        self.data_derQ=[]
        self.data_dersigma=[]
        self.data_epsilons=[]
        self.data2_Q=[]
        self.data2_sigma=[]
        self.data2_derQ=[]
        self.data2_dersigma=[]
        self.data2_epsilons=[]

        self.datafilename1=''
        self.datafilename2=''

        self.weekendfactora=[0.0,0.0] # first factor for weekend columns, second for rows
        self.weekendfactorb=[0.0,0.0]

        # init values
        self.ix=-1
        self.iy=-1
        self.startix=-1
        self.startiy=-1

        # panel buffers
        self.panelbuffers={}
        self.panelredrawstatus={}

        self.datetimeselectionsstart=['']
        self.datetimeselectionsend=['']
        self.datetimeselectionsstart2=['']
        self.datetimeselectionsend2=['']

        # other init
        self.abcurvepanelsize='small'
        self.markercolsa=[]
        self.markercolsb=[]

        # init pygame
        pygame.init()

        # create window
        (self.xsize,self.ysize)=wx.GetDisplaySize()
        self.ysize-=48 # windows bar
        wx.Frame.__init__(self, None, wx.ID_ANY, self.appname,
            pos=(0, 0), size=(self.xsize, self.ysize))
            #pos=(100, 100), size=(self.xsize, self.ysize))
        self.SetBackgroundColour("white")

        self.topsizer=wx.BoxSizer(wx.VERTICAL)
        set1sizer=wx.BoxSizer(wx.VERTICAL)
        self.set1l1sizer=wx.BoxSizer(wx.HORIZONTAL)
        set1l2sizer=wx.BoxSizer(wx.HORIZONTAL)
        set1l3sizer=wx.BoxSizer(wx.HORIZONTAL)
        self.set1l4sizer=wx.BoxSizer(wx.HORIZONTAL)
        set2sizer=wx.BoxSizer(wx.VERTICAL)
        self.set2l1sizer=wx.BoxSizer(wx.HORIZONTAL)
        set2l2sizer=wx.BoxSizer(wx.HORIZONTAL)
        set2l3sizer=wx.BoxSizer(wx.HORIZONTAL)
        self.set2l4sizer=wx.BoxSizer(wx.HORIZONTAL)
        botsizer=wx.BoxSizer(wx.HORIZONTAL)
        
        self.topsizer.Add(set1sizer, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        self.topsizer.Add(wx.StaticLine(self), proportion=0, flag=wx.ALL|wx.EXPAND, border=3)
        self.topsizer.Add(set2sizer, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        self.topsizer.Add(wx.StaticLine(self), proportion=0, flag=wx.ALL|wx.EXPAND, border=3)
        self.topsizer.Add(botsizer, proportion=0, flag=wx.ALL|wx.EXPAND, border=3)
        
        set1sizer.Add(self.set1l1sizer, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        set1sizer.Add(set1l2sizer, proportion=0, flag=wx.ALL, border=3)
        set1sizer.Add(set1l3sizer, proportion=0, flag=wx.ALL, border=3)
        set1sizer.Add(self.set1l4sizer, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        set2sizer.Add(self.set2l1sizer, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        set2sizer.Add(set2l2sizer, proportion=0, flag=wx.ALL, border=3)
        set2sizer.Add(set2l3sizer, proportion=0, flag=wx.ALL, border=3)
        set2sizer.Add(self.set2l4sizer, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        
        # graph panels measured pattern
        self.fppanel = wx.Panel(self)
        self.set1l1sizer.Add(self.fppanel, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        self.fppanelstaticbitmap = wx.StaticBitmap(self.fppanel)
        self.panelredrawstatus["fppanel"]=True
        self.Refresh()        

        # main buttons

        loadfilebutton=wx.Button(self, id=-1, label='Load file')
        loadfilebutton.SetFont(self.textfont)
        set1l2sizer.Add(loadfilebutton, proportion=0, flag=wx.ALL, border=3)

        loadfilebutton.Bind(wx.EVT_BUTTON, lambda event: self.butloadfile(wx.EVT_BUTTON,1))
        loadfilebutton.SetToolTip(wx.ToolTip("Load flow measurement time series"))

        label=wx.StaticText(self,-1,'Period Window:')
        label.SetFont(self.textfont)
        set1l2sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)

        self.minxfield=wx.ComboBox(self, -1, value='',choices=self.datetimeselectionsstart, size=(200,20))
        self.maxxfield=wx.ComboBox(self, -1, value='',choices=self.datetimeselectionsend,size=(200,20))
        self.minxfield.SetValue(conversion.exceldate2string(min(self.datx)))
        self.maxxfield.SetValue(conversion.exceldate2string(max(self.datx)))
        self.minxfield.SetFont(self.textfont)
        self.maxxfield.SetFont(self.textfont)
        set1l2sizer.Add(self.minxfield, proportion=0, flag=wx.ALL, border=3)
        set1l2sizer.Add(self.maxxfield, proportion=0, flag=wx.ALL, border=3)

        zoombutton=wx.Button(self, id=-1, label='Zoom')
        zoombutton.SetFont(self.textfont)
        zoombutton.Bind(wx.EVT_BUTTON, self.butzoom)
        zoombutton.SetToolTip(wx.ToolTip("Zoom in to specified time window"))
        set1l2sizer.Add(zoombutton, proportion=0, flag=wx.ALL, border=3)


        scaleshiftbutton=wx.Button(self, id=-1, label='Gen./Scale/Shift')
        scaleshiftbutton.SetFont(self.textfont)
        scaleshiftbutton.Bind(wx.EVT_BUTTON, lambda event: self.createoperationswindow(wx.EVT_BUTTON,self.title1.GetValue(),1))
        scaleshiftbutton.SetToolTip(wx.ToolTip("Generate, scale and/or shift flow pattern"))
        set1l2sizer.Add(scaleshiftbutton, proportion=0, flag=wx.ALL, border=3)

        aggregatebutton=wx.Button(self, id=-1, label='Aggregate')
        aggregatebutton.SetFont(self.textfont)
        aggregatebutton.Bind(wx.EVT_BUTTON, lambda event: self.aggregate(wx.EVT_BUTTON,self.title1.GetValue(),1))
        aggregatebutton.SetToolTip(wx.ToolTip("Aggregate data in time blocks of arbitrary size."))
        set1l2sizer.Add(aggregatebutton, proportion=0, flag=wx.ALL, border=3)

        savedata1button=wx.Button(self, id=-1, label='Save')
        savedata1button.SetFont(self.textfont)
        savedata1button.Bind(wx.EVT_BUTTON, lambda event: self.presssavedata(wx.EVT_BUTTON,1))
        savedata1button.SetToolTip(wx.ToolTip("Save modified time series (zoom)."))
        set1l2sizer.Add(savedata1button, proportion=0, flag=wx.ALL, border=3)


        #  time window
        label=wx.StaticText(self,-1,'Time window:')
        label.SetFont(self.textfont)
        set1l3sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)

        self.mintfield=wx.TextCtrl(self)
        set1l3sizer.Add(self.mintfield, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        self.maxtfield=wx.TextCtrl(self)
        set1l3sizer.Add(self.maxtfield, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        self.mintfield.SetValue("00:00:00")
        self.maxtfield.SetValue("24:00:00")
        self.mintfield.SetFont(self.textfont)
        self.maxtfield.SetFont(self.textfont)
        label=wx.StaticText(self,-1,'Value range:')
        label.SetFont(self.textfont)
        self.minyfield=wx.TextCtrl(self)
        self.maxyfield=wx.TextCtrl(self)
        self.minyfield.SetValue("0.0")
        self.maxyfield.SetValue("500.0")
        self.minyfield.SetFont(self.textfont)
        self.maxyfield.SetFont(self.textfont)
        set1l3sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)
        set1l3sizer.Add(self.minyfield, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        set1l3sizer.Add(self.maxyfield, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)

        label=wx.StaticText(self,-1,'Title:')
        label.SetFont(self.textfont)
        set1l3sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)
        self.title1=wx.TextCtrl(self)
        self.title1.SetFont(self.textfont)
        self.title1.SetValue('dataset 1')
        set1l3sizer.Add(self.title1, proportion=2, flag=wx.ALL|wx.EXPAND, border=3)

        self.crosstabwinbut=wx.Button(self, id=-1, label='CFPD block')
        self.crosstabwinbut.SetFont(self.textfont)
        self.crosstabwinbut.Bind(wx.EVT_BUTTON, lambda event: self.createcrosstabwin(wx.EVT_BUTTON,1))
        self.crosstabwinbut.SetToolTip(wx.ToolTip("Perform CFPD block analysis for this dataset"))
        set1l3sizer.Add(self.crosstabwinbut, proportion=0, flag=wx.ALL, border=3)


        # zoom panels
        self.zoompansizer=wx.BoxSizer(wx.HORIZONTAL)
        self.zoompanelysize=self.fppanelysize
        self.zoompanel = wx.Panel(self)
        self.panelredrawstatus["zoompanel"]=True
        self.zoompanelstaticbitmap = wx.StaticBitmap(self.zoompanel)
        self.onzoompaint(wx.EVT_PAINT,self.zoompanel,"zoompanel")
        self.zoompansizer.Add(self.zoompanel, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)




        self.histpansizer=wx.BoxSizer(wx.HORIZONTAL)
        self.histpanel = wx.Panel(self)
        self.panelredrawstatus["histpanel"]=True
        self.histpanelstaticbitmap = wx.StaticBitmap(self.histpanel)
        self.onhistpaint(wx.EVT_PAINT,self.histpanel, "histpanel")
        self.histpansizer.Add(self.histpanel, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        self.set1l4sizer.Add(self.zoompansizer, proportion=6, flag=wx.ALL|wx.EXPAND, border=3)
        self.set1l4sizer.Add(self.histpansizer, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)




        # graph panels pattern 2

        self.fppanel2 = wx.Panel(self)
        self.set2l1sizer.Add(self.fppanel2, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)

        self.panelredrawstatus["fppanel2"]=True
        self.fppanel2staticbitmap = wx.StaticBitmap(self.fppanel2)
        self.onfppaint2(wx.EVT_PAINT,self.fppanel2,"fppanel2")


        # main buttons
        loadfilebutton2=wx.Button(self, id=-1, label='Load file')
        loadfilebutton2.SetFont(self.textfont)
        loadfilebutton2.Bind(wx.EVT_BUTTON, lambda event: self.butloadfile(wx.EVT_BUTTON,2))
        loadfilebutton2.SetToolTip(wx.ToolTip("Load flow measurement time series"))
        set2l2sizer.Add(loadfilebutton2, proportion=0, flag=wx.ALL, border=3)


        label=wx.StaticText(self,-1,'Period Window:')
        label.SetFont(self.textfont)
        set2l2sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)
        self.minxfield2=wx.ComboBox(self, -1, value='',choices=self.datetimeselectionsstart, size=(200,20))
        self.maxxfield2=wx.ComboBox(self, -1, value='',choices=self.datetimeselectionsend, size=(200,20))
        self.minxfield2.SetValue(conversion.exceldate2string(min(self.datx)))
        self.maxxfield2.SetValue(conversion.exceldate2string(max(self.datx)))
        self.minxfield2.SetFont(self.textfont)
        self.maxxfield2.SetFont(self.textfont)
        set2l2sizer.Add(self.minxfield2, proportion=0, flag=wx.ALL, border=3)
        set2l2sizer.Add(self.maxxfield2, proportion=0, flag=wx.ALL, border=3)

        zoombutton2=wx.Button(self, id=-1, label='Zoom')
        zoombutton2.SetFont(self.textfont)
        zoombutton2.Bind(wx.EVT_BUTTON, self.butzoom2)
        zoombutton2.SetToolTip(wx.ToolTip("Zoom in to specified time window"))
        set2l2sizer.Add(zoombutton2, proportion=0, flag=wx.ALL, border=3)

        scaleshiftbutton2=wx.Button(self, id=-1, label='Gen./Scale/Shift')
        scaleshiftbutton2.SetFont(self.textfont)
        scaleshiftbutton2.Bind(wx.EVT_BUTTON, lambda event: self.createoperationswindow(wx.EVT_BUTTON,self.title2.GetValue(),2))
        scaleshiftbutton2.SetToolTip(wx.ToolTip("Scale and/or shift flow pattern"))
        set2l2sizer.Add(scaleshiftbutton2, proportion=0, flag=wx.ALL, border=3)

        aggregatebutton2=wx.Button(self, id=-1, label='Aggregate')
        aggregatebutton2.SetFont(self.textfont)
        aggregatebutton2.Bind(wx.EVT_BUTTON, lambda event: self.aggregate(wx.EVT_BUTTON,self.title1.GetValue(),2))
        aggregatebutton2.SetToolTip(wx.ToolTip("Aggregate data in time blocks of arbitrary size."))
        set2l2sizer.Add(aggregatebutton2, proportion=0, flag=wx.ALL, border=3)

        savedata2button=wx.Button(self, id=-1, label='Save')
        savedata2button.SetFont(self.textfont)
        savedata2button.Bind(wx.EVT_BUTTON, lambda event: self.presssavedata(wx.EVT_BUTTON,2))
        savedata2button.SetToolTip(wx.ToolTip("Save modified time series (zoom)."))
        set2l2sizer.Add(savedata2button, proportion=0, flag=wx.ALL, border=3)



        #  time window
        label=wx.StaticText(self,-1,'Time window:')
        label.SetFont(self.textfont)
        set2l3sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)
        self.mintfield2=wx.TextCtrl(self)
        self.maxtfield2=wx.TextCtrl(self)
        self.mintfield2.SetValue("00:00:00")
        self.maxtfield2.SetValue("24:00:00")
        self.mintfield2.SetFont(self.textfont)
        self.maxtfield2.SetFont(self.textfont)
        label=wx.StaticText(self,-1,'Value range:')
        label.SetFont(self.textfont)
        self.minyfield2=wx.TextCtrl(self)
        self.maxyfield2=wx.TextCtrl(self)
        self.minyfield2.SetValue("0.0")
        self.maxyfield2.SetValue("500.0")
        self.minyfield2.SetFont(self.textfont)
        self.maxyfield2.SetFont(self.textfont)
        set2l3sizer.Add(self.mintfield2, proportion=0, flag=wx.ALL, border=3)
        set2l3sizer.Add(self.maxtfield2, proportion=0, flag=wx.ALL, border=3)
        set2l3sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)
        set2l3sizer.Add(self.minyfield2, proportion=0, flag=wx.ALL, border=3)
        set2l3sizer.Add(self.maxyfield2, proportion=0, flag=wx.ALL, border=3)
        
        label=wx.StaticText(self,-1,'Title:')
        label.SetFont(self.textfont)
        set2l3sizer.Add(label, proportion=0, flag=wx.ALL|wx.ALIGN_CENTER_VERTICAL, border=3)
        self.title2=wx.TextCtrl(self)
        self.title2.SetFont(self.textfont)
        self.title2.SetValue('dataset 2')
        set2l3sizer.Add(self.title2, proportion=2, flag=wx.ALL|wx.EXPAND, border=3)


        self.zoompansizer2=wx.BoxSizer(wx.HORIZONTAL)
        self.zoompanel2 = wx.Panel(self)
        self.panelredrawstatus["zoompanel2"]=True
        self.zoompanel2staticbitmap = wx.StaticBitmap(self.zoompanel2)
        self.onzoompaint2(wx.EVT_PAINT, self.zoompanel2, "zoompanel2")
        self.zoompansizer2.Add(self.zoompanel2, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)

        self.histpansizer2=wx.BoxSizer(wx.HORIZONTAL)
        self.histpanel2 = wx.Panel(self)
        self.panelredrawstatus["histpanel2"]=True
        self.histpanel2staticbitmap = wx.StaticBitmap(self.histpanel2)
        self.onhistpaint2(wx.EVT_PAINT,self.histpanel2,"histpanel2")
        self.histpansizer2.Add(self.histpanel2, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)

        self.set2l4sizer.Add(self.zoompansizer2, proportion=6, flag=wx.ALL|wx.EXPAND, border=3)
        self.set2l4sizer.Add(self.histpansizer2, proportion=1, flag=wx.ALL|wx.EXPAND, border=3)
        # analysis control

        botsizer.Add(wx.BoxSizer(wx.HORIZONTAL), proportion=1, flag=wx.ALL|wx.EXPAND, border=3)

        self.jppwinbut=wx.Button(self, id=-1, label='CFPD')
        self.jppwinbut.SetFont(self.textfont)
        self.jppwinbut.Bind(wx.EVT_BUTTON, self.calcjpp)
        self.jppwinbut.SetToolTip(wx.ToolTip("Perform single CFPD analysis on data sets"))
        botsizer.Add(self.jppwinbut, proportion=0, flag=wx.ALL, border=3)

        self.infowinbut=wx.Button(self, id=-1, label='Info')
        self.infowinbut.SetFont(self.textfont)
        self.infowinbut.Bind(wx.EVT_BUTTON, self.infowin)
        self.infowinbut.SetToolTip(wx.ToolTip("Show info window"))
        botsizer.Add(self.infowinbut, proportion=0, flag=wx.ALL, border=3)



        self.SetSizer(self.topsizer)

        # show window
        self.Show(True)
        self.Bind(wx.EVT_SIZE, self.onSize)
        
        # redraw panels with correct size
        self.updateBitmaps(wx.EVT_SIZE)
        
    def onSize(self,evt):
        self.Layout()
        self.updateBitmaps(evt)
        
    def updateBitmaps(self,evt):
        # redraw panels with correct size
        self.onfppaint(wx.EVT_PAINT,self.fppanel,"fppanel",forceredraw=True )
        self.onfppaint2(wx.EVT_PAINT,self.fppanel2,"fppanel2",forceredraw=True )
        self.onzoompaint(wx.EVT_PAINT,self.zoompanel,"zoompanel",forceredraw=True )
        self.onzoompaint2(wx.EVT_PAINT,self.zoompanel2,"zoompanel2",forceredraw=True )
        self.onhistpaint(wx.EVT_PAINT,self.histpanel,"histpanel",forceredraw=True )
        self.onhistpaint2(wx.EVT_PAINT,self.histpanel2,"histpanel2",forceredraw=True )        

    def onfppaint(self, evt,panel,panelname,forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
           (xl,yl)=panel.Size
           self.fppanelstaticbitmap.SetClientSize(xl,yl)

           self.panelbuffers[panelname]=visual.drawcurves(self.fppanelstaticbitmap,[[self.datx,self.daty]],[0],True,[-1,-1],[0,-1],xl,yl, True, [self.minxzoom, self.maxxzoom], levels=True)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(panel,self.panelbuffers[panelname])
           
        self.fppanelstaticbitmap.SetBitmap(self.panelbuffers[panelname])
      
    def onfppaint2(self, evt, panel,panelname,forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
           (xl,yl)=panel.Size
           self.fppanel2staticbitmap.SetClientSize(xl,yl)
           self.panelbuffers[panelname]=visual.drawcurves(self.fppanel2staticbitmap,[[self.datx2,self.daty2]],[0],True,[-1,-1],[0,-1],xl,yl, True, [self.minxzoom2, self.maxxzoom2], levels=True)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(panel,self.panelbuffers[panelname])
           
        self.fppanel2staticbitmap.SetBitmap(self.panelbuffers[panelname])


    def onzoompaint(self, evt, panel,panelname,forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
           try:
              miny=float(self.minyfield.GetValue())
           except: 
              miny=-1
           try:
              maxy=float(self.maxyfield.GetValue())
           except: 
              miny=-1
           (xl,yl)=panel.Size
           self.zoompanelstaticbitmap.SetClientSize(xl,yl)
           self.panelbuffers[panelname]=visual.drawcurves(self.zoompanelstaticbitmap,[[self.datx,self.daty]],[0],True,[self.minxzoom, self.maxxzoom],[miny,maxy],xl,yl, False, [0,0], levels=True)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(panel,self.panelbuffers[panelname])
           
        self.zoompanelstaticbitmap.SetBitmap(self.panelbuffers[panelname])


    def onzoompaint2(self, evt, panel,panelname,forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
           try:
              miny=float(self.minyfield2.GetValue())
           except: 
              miny=-1
           try:
              maxy=float(self.maxyfield2.GetValue())
           except: 
              miny=-1
           (xl,yl)=panel.Size
           self.zoompanel2staticbitmap.SetClientSize(xl,yl)
           self.panelbuffers[panelname]=visual.drawcurves(self.zoompanel2staticbitmap,[[self.datx2,self.daty2]],[0],True,[self.minxzoom2, self.maxxzoom2],[miny,maxy],xl,yl, False, [0,0], levels=True)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(panel,self.panelbuffers[panelname])
           
           
        self.zoompanel2staticbitmap.SetBitmap(self.panelbuffers[panelname])

      
    def onhistpaint(self,evt, panel,panelname,equal=False,forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
        #try:
           miny=float(self.minyfield.GetValue())
           maxy=float(self.maxyfield.GetValue())
           mint=conversion.timstr2num(self.mintfield.GetValue())
           maxt=conversion.timstr2num(self.maxtfield.GetValue())

           [sx,sy,si]=mystat.subset(self.datx, self.daty, self.minxzoom, self.maxxzoom, miny, maxy, mint, maxt)

           if equal:
              miny2=float(self.minyfield2.GetValue())
              maxy2=float(self.maxyfield2.GetValue())
              mint2=conversion.timstr2num(self.mintfield2.GetValue())
              maxt2=conversion.timstr2num(self.maxtfield2.GetValue())
              [sx2,sy2,si2]=mystat.subset(self.datx2, self.daty2, self.minxzoom2, self.maxxzoom2, miny2, maxy2, mint2, maxt2)
              minsy=min(min(sy),min(sy2))
              maxsy=max(max(sy),max(sy2))

              if (len(sy)>0):
                 nbin=int(math.sqrt(float(len(sy))))
                 nbin=min(nbin,50) # max 50 bins
                 [hist,bounds]=numpy.histogram(sy,nbin,range=(minsy,maxsy))
              else:
                 hist=[]
                 bounds=[0,1]

           else:
              if (len(sy)>0):
                 nbin=int(math.sqrt(float(len(sy))))
                 nbin=min(nbin,50) # max 50 bins
                 [hist,bounds]=numpy.histogram(sy,nbin)
              else:
                 hist=[]
                 bounds=[0,1]
 
           (xl,yl)=panel.Size
           self.histpanelstaticbitmap.SetClientSize(xl,yl)           
           self.panelbuffers[panelname]=visual.drawhist(self.histpanelstaticbitmap,hist,bounds,xl,yl,values=True)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(panel,self.panelbuffers[panelname])
           
        self.histpanelstaticbitmap.SetBitmap(self.panelbuffers[panelname])


        #except:
           #dlg = wx.MessageDialog(self, 'Histogram error', 'Error', wx.ICON_ERROR)
           #dlg.ShowModal()

    def onhistpaint2(self,evt, panel,panelname,equal=False, forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
        #try:
           miny=float(self.minyfield2.GetValue())
           maxy=float(self.maxyfield2.GetValue())
           mint=conversion.timstr2num(self.mintfield2.GetValue())
           maxt=conversion.timstr2num(self.maxtfield2.GetValue())

           [sx,sy,si]=mystat.subset(self.datx2, self.daty2, self.minxzoom2, self.maxxzoom2, miny, maxy, mint, maxt)



           if equal:
              miny2=float(self.minyfield.GetValue())
              maxy2=float(self.maxyfield.GetValue())
              mint2=conversion.timstr2num(self.mintfield.GetValue())
              maxt2=conversion.timstr2num(self.maxtfield.GetValue())
              [sx2,sy2,si2]=mystat.subset(self.datx, self.daty, self.minxzoom, self.maxxzoom, miny2, maxy2, mint2, maxt2)
              minsy=min(min(sy),min(sy2))
              maxsy=max(max(sy),max(sy2))

              if (len(sy)>0):
                 nbin=int(math.sqrt(float(len(sy))))
                 nbin=min(nbin,50) # max 50 bins
                 [hist,bounds]=numpy.histogram(sy,nbin,range=(minsy,maxsy))
              else:
                 hist=[]
                 bounds=[0,1]
           else:
              if (len(sy)>0):
                 nbin=int(math.sqrt(float(len(sy))))
                 nbin=min(nbin,50) # max 50 bins
                 [hist,bounds]=numpy.histogram(sy,nbin)
              else:
                 hist=[]
                 bounds=[0,1]


 
           (xl,yl)=panel.Size
           self.histpanel2staticbitmap.SetClientSize(xl,yl)  
           self.panelbuffers[panelname]=visual.drawhist(self.histpanel2staticbitmap,hist,bounds,xl,yl,values=True)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(panel,self.panelbuffers[panelname])
           
        self.histpanel2staticbitmap.SetBitmap(self.panelbuffers[panelname])


    def ongp12paint(self,evt,forceredraw=False):
        self.ongp1paint(wx.EVT_PAINT,"graphpanel1",forceredraw=forceredraw)
        self.ongp2paint(wx.EVT_PAINT,"graphpanel2",forceredraw=forceredraw)

    def ongp1paint(self,evt,panelname,forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:

           if (len(self.daty)==2):
              return
   
           if len(self.data_epsilons)>0:
              dminepsilon=min(self.data_epsilons)
              dmaxepsilon=max(self.data_epsilons)
           else:
              dminepsilon=-2
              dmaxepsilon=2
           if len(self.data2_epsilons)>0:
              bminepsilon=min(self.data2_epsilons)
              bmaxepsilon=max(self.data2_epsilons)
           else:
              bminepsilon=-2
              bmaxepsilon=2
   
           minepsilon=max(-3,min(bminepsilon,dminepsilon))
           maxepsilon=min(bmaxepsilon,dmaxepsilon)
   
           syn1=self.showsynth1.GetValue() 
           syn2=self.showsynth2.GetValue() 
   
           curves=[[self.data_epsilons,self.data_Q],[self.data_epsilons,self.data_derQ],[self.data2_epsilons,self.data2_Q],[self.data2_epsilons,self.data2_derQ]]
           colors=[0,1,2,3]
   
           t1=self.title1.GetValue()
           t2=self.title2.GetValue()
           legend=[t1+': Q',t1+': dQ/de',t2+': Q',t2+': dQ/de']
   
           icol=3
           if syn1:
              curves.append([self.syneps1,self.synE1])
              curves.append([self.syneps1,self.dersynE1])
              icol+=1
              colors.append(icol)
              icol+=1
              colors.append(icol)
              legend.append(t1+': synthetic Q')
              legend.append(t1+': synthetic dQ/de')
           if syn2:
              curves.append([self.syneps2,self.synE2])
              curves.append([self.syneps2,self.dersynE2])
              icol+=1
              colors.append(icol)
              icol+=1
              colors.append(icol)
              legend.append(t2+': synthetic Q')
              legend.append(t2+': synthetic dQ/de')
              
           self.panelbuffers[panelname]=visual.drawcurves(self.graphpanel1,curves,colors,False,[minepsilon,maxepsilon],[-1,-1],self.gpxsize,self.gpysize, False, [0,0], legend, levels=True, vergrid=True, linethickness=2,fontsize=14)
           self.panelredrawstatus[panelname]=False
   
        else:
           visual.drawBitmap(self.graphpanel1,self.panelbuffers[panelname])
        

    def ongp2paint(self,evt,panelname, forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
           if (len(self.daty2)==2):
              return
   
           if len(self.data_epsilons)>0:
              dminepsilon=min(self.data_epsilons)
              dmaxepsilon=max(self.data_epsilons)
           else:
              dminepsilon=-2
              dmaxepsilon=2
           if len(self.data2_epsilons)>0:
              bminepsilon=min(self.data2_epsilons)
              bmaxepsilon=max(self.data2_epsilons)
           else:
              bminepsilon=-2
              bmaxepsilon=2
   
           minepsilon=max(-3,min(bminepsilon,dminepsilon))
           maxepsilon=min(bmaxepsilon,dmaxepsilon)
   
           syn1=self.showsynth1.GetValue() 
           syn2=self.showsynth2.GetValue() 
   
           curves=[[self.data_epsilons,self.data_sigma],[self.data_epsilons,self.data_dersigma],[self.data2_epsilons,self.data2_sigma],[self.data2_epsilons,self.data2_dersigma]]
           colors=[0,1,2,3]
   
           t1=self.title1.GetValue()
           t2=self.title2.GetValue()
           legend=[t1+': s',t1+': ds/de',t2+': s',t2+': ds/de']
   
           icol=3
           if syn1:
              curves.append([self.syneps1,self.synS1])
              curves.append([self.syneps1,self.dersynS1])
              icol+=1
              colors.append(icol)
              icol+=1
              colors.append(icol)
              legend.append(t1+': synthetic s')
              legend.append(t1+': synthetic ds/de')
           if syn2:
              curves.append([self.syneps2,self.synS2])
              curves.append([self.syneps2,self.dersynS2])
              icol+=1
              colors.append(icol)
              icol+=1
              colors.append(icol)
              legend.append(t2+': synthetic s')
              legend.append(t2+': synthetic ds/de')
              
           self.panelbuffers[panelname]=visual.drawcurves(self.graphpanel2,curves,colors,False,[minepsilon,maxepsilon],[-1,-1],self.gpxsize,self.gpysize, False, [0,0], legend, levels=True, vergrid=True, linethickness=2,fontsize=14)
           self.panelredrawstatus[panelname]=False

        else:
           visual.drawBitmap(self.graphpanel2,self.panelbuffers[panelname])

    def butzoom(self,evt):

        minxstr=self.minxfield.GetValue()
        maxxstr=self.maxxfield.GetValue()
       
        try:
           s1=conversion.string2exceldate(minxstr)
           s2=conversion.string2exceldate(maxxstr)
           self.minxzoom=s1
           self.maxxzoom=s2

           self.panelredrawstatus["fppanel"]=True
           self.onfppaint(wx.EVT_PAINT,self.fppanel,"fppanel")
           self.panelredrawstatus["zoompanel"]=True
           self.panelredrawstatus["histpanel"]=True
           self.onzoompaint(wx.EVT_PAINT,self.zoompanel,"zoompanel")
           self.onhistpaint(wx.EVT_PAINT, self.histpanel,"histpanel")

           base=self.title1.GetValue().split(' ')[0]
           self.setlabel1(base)

        except:
           dlg = wx.MessageDialog(self, 'Invalid time window', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()

    def butzoom2(self,evt):

        minxstr=self.minxfield2.GetValue()
        maxxstr=self.maxxfield2.GetValue()
       
        try:
           s1=conversion.string2exceldate(minxstr)
           s2=conversion.string2exceldate(maxxstr)
           self.minxzoom2=s1
           self.maxxzoom2=s2
           #print 'butzoom range '+str([s1,s2])

           self.panelredrawstatus["fppanel2"]=True
           self.onfppaint2(wx.EVT_PAINT, self.fppanel2,"fppanel2")
           self.panelredrawstatus["zoompanel2"]=True
           self.panelredrawstatus["histpanel2"]=True
           self.onzoompaint2(wx.EVT_PAINT, self.zoompanel2,"zoompanel2")
           self.onhistpaint2(wx.EVT_PAINT, self.histpanel2,"histpanel2")

           base=self.title2.GetValue().split(' ')[0]
           self.setlabel2(base)

        except:
           dlg = wx.MessageDialog(self, 'Invalid time window', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()
           
           
    def butloadfile(self,evt,ifile):
    
        path=''
        dlg = wx.FileDialog(self,message='Load data file '+str(ifile), defaultDir=self.cwd, wildcard="All compatible files|*.csv;*.xls|Comma Separated Values (*.csv)|*.csv|Excel -2003 files (*.xls)|*.xls", style=wx.FD_OPEN)
        if dlg.ShowModal() == wx.ID_OK:
           path=dlg.GetPath()
           self.cwd=fsconfig.strippath(path)
           fsconfig.saveconfig([self.cwd,self.appsize])
        dlg.Destroy()
        if (path==''):
           dlg = wx.MessageDialog(self, 'No file selected', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()
        else:
           #try:

              ext=path.split('.')[-1] # file extension

              print('path='+path)
              print('ext='+ext)
 
              if ext=='csv':
                 # csv file
                 #f=open(path,'r')
                 #dat=f.readlines()
                 #f.close()
                 #[ldatx, ldaty, ldatt]=self.parsefile(dat)
                 [success, ldatx, ldaty, ldatt]=self.loadandparsecsv(path)
              elif ext=='xls':
                 # excel file
                 [success, ldatx, ldaty, ldatt]=self.loadandparsexls(path)

                
                 if not success:
                    raise Exception('Failed to load xls file')
              elif ext=='xlsx':
                 # excel file
                 [success, ldatx, ldaty, ldatt]=self.loadandparsexlsx(path) # make based on openpyxl
                
                 if not success:
                    raise Exception('Failed to load xlsx file')
              else:
                 raise Exception('extension')
   
              if ifile==1:
                 self.datafilename1=path
                 self.datx=ldatx
                 self.daty=ldaty
                 self.datt=ldatt

                 # update date combobox options list
                 self.datetimeselectionsstart,self.datetimeselectionsend=supportfunctions.getDateTimeSelections(self.datx)
                 self.minxfield.Clear() # remove combobox options
                 for item in self.datetimeselectionsstart:
                    self.minxfield.Append(item)
                 self.maxxfield.Clear() # remove combobox options
                 for item in self.datetimeselectionsend:
                    self.maxxfield.Append(item)
                    
                 v1=conversion.exceldate2string(min(self.datx))
                 v2=conversion.exceldate2string(max(self.datx))
                 self.minxfield.SetValue(v1)
                 self.maxxfield.SetValue(v2)
                 
               
                 self.maxyfield.SetValue(str(max(self.daty)))
                 self.minxzoom=min(self.datx)
                 self.maxxzoom=max(self.datx)
                 self.panelredrawstatus["fppanel"]=True
                 self.panelredrawstatus["zoompanel"]=True
                 self.panelredrawstatus["histpanel"]=True
                 self.onfppaint(wx.EVT_PAINT,self.fppanel,"fppanel")
                 self.onzoompaint(wx.EVT_PAINT, self.zoompanel,"zoompanel")
                 self.onhistpaint(wx.EVT_PAINT, self.histpanel,"histpanel")


              elif ifile==2:
                 self.datafilename2=path
                 self.datx2=ldatx
                 self.daty2=ldaty
                 self.datt2=ldatt

                 # update date combobox options list
                 self.datetimeselectionsstart2,self.datetimeselectionsend2=supportfunctions.getDateTimeSelections(self.datx2)
                 self.minxfield2.Clear() # remove combobox options
                 for item in self.datetimeselectionsstart2:
                    self.minxfield2.Append(item)
                 self.maxxfield2.Clear() # remove combobox options
                 for item in self.datetimeselectionsend2:
                    self.maxxfield2.Append(item)
                 self.minxfield2.SetValue(conversion.exceldate2string(min(self.datx2)))
                 self.maxxfield2.SetValue(conversion.exceldate2string(max(self.datx2)))
                 self.maxyfield2.SetValue(str(max(self.daty2)))
                 self.minxzoom2=min(self.datx2)
                 self.maxxzoom2=max(self.datx2)
                 self.panelredrawstatus["fppanel2"]=True
                 self.panelredrawstatus["zoompanel2"]=True
                 self.panelredrawstatus["histpanel2"]=True
                 self.onfppaint2(wx.EVT_PAINT,self.fppanel2, "fppanel2")
                 self.onzoompaint2(wx.EVT_PAINT, self.zoompanel2, "zoompanel2")
                 self.onhistpaint2(wx.EVT_PAINT, self.histpanel2, "histpanel2")

              else:
                 print('error ifile='+str(ifile))
                 sys.exit(1)
   
              ltit=path.split('/')[-1].replace('.'+ext,'')
              if ifile==1:
                 self.setlabel1(ltit)
              else:
                 self.setlabel2(ltit)
   
              dlg = wx.MessageDialog(self, 'Data file '+str(ifile)+' loaded', 'Succes', wx.ICON_INFORMATION)
              dlg.ShowModal()
              self.butzoom(wx.EVT_BUTTON)
           #except:
              #dlg = wx.MessageDialog(self, 'Data file '+str(ifile)+' error', 'Error', wx.ICON_ERROR)
              #dlg.ShowModal()
  

    def setlabel1(self,base):
        # set title

             

        d1=conversion.exceldate2stringd(conversion.string2exceldate(self.minxfield.GetValue()))
        if (max(self.datx)%1<1e-2):
           d2=conversion.exceldate2stringd(conversion.string2exceldate(self.maxxfield.GetValue())-1)
        else:
           d2=conversion.exceldate2stringd(conversion.string2exceldate(self.maxxfield.GetValue()))
        if d1==d2:
           ltit=base+' '+d1
        else:
           ltit=base+' '+d1+' tm '+d2
        self.title1.SetValue(ltit)


    def setlabel2(self,base):
        # set title
        d1=conversion.exceldate2stringd(conversion.string2exceldate(self.minxfield2.GetValue()))
        if (max(self.datx2)%1<1e-2):
           d2=conversion.exceldate2stringd(conversion.string2exceldate(self.maxxfield2.GetValue())-1)
        else:
           d2=conversion.exceldate2stringd(conversion.string2exceldate(self.maxxfield2.GetValue()))
        if d1==d2:
           ltit=base+' '+d1
        else:
           ltit=base+' '+d1+' tm '+d2
        self.title2.SetValue(ltit)

    def loadandparsexlsx(self,path):
           book=openpyxl.load_workbook(path)
           sheets=book.get_sheet_names()
           choices=[]
           for sheet in sheets:
              choices.append(str(sheet))

           # select sheet
           dlg = wx.SingleChoiceDialog(self,'Worksheets', 'Select worksheet', choices)
           if dlg.ShowModal() != wx.ID_OK:
              return False,[],[],[]
           selectedsheet=dlg.GetSelection() # sheet index
           dlg.Destroy()
           sheet=book.get_sheet_by_name(str(selectedsheet))

           # select column
           selections=[]
           for col_index in range(0,sheet.ncols):
              s=str(col_index)+':"'+str(sheet.cell(row=0,column=col_index))+'"' # include first line value in selections
              selections.append(s)
           dlg = wx.SingleChoiceDialog(self,'Columns', 'Select timestamp column', selections)
           if dlg.ShowModal() != wx.ID_OK:
              return False,[],[],[]
           tscolumn=dlg.GetSelection() # not string but index
           dlg.Destroy()
           dlg = wx.SingleChoiceDialog(self,'Columns', 'Select data column', selections)
           if dlg.ShowModal() != wx.ID_OK:
              print('dlg.ShowModal()='+str(dlg.ShowModal()))
              return False,[],[],[]
           datacolumn=dlg.GetSelection() # not string but index
           dlg.Destroy()
           dlg = wx.MessageDialog(parent=self,caption='Rows', message='Skip first row?',style=wx.YES|wx.NO)
           if dlg.ShowModal() == wx.ID_YES:
              istart=1
           else:
              istart=0

           dlg.Destroy()

           # fill arrays
           datx=[]
           daty=[]
           datt=[]
           for row_index in range(istart,sheet.nrows):
              x=sheet.cell(row = row_index, column = tscolumn)
              y=sheet.cell(row = row_index, column = datacolumn)
              #print 'x,y='+str([x,y])
              if not (x=='' or y==''):
                 try:
                    lx=float(x)
                    ly=float(y)
                    lt=float(lx)%1.0
                    datx.append(lx)
                    daty.append(ly)
                    datt.append(lt)
                 except:
                    pass
                    #print 'skipping row '+str(row_index)+' (not parsable)'
              else:
                 pass
                 #print 'skipping row '+str(row_index)+' (empty)'



           return True, datx, daty, datt



    def loadandparsexlsheader(self,path):
        book=xlrd.open_workbook(path)
        sheet=book.sheet_by_index(0)
        labels=[]
        for col_index in range(0,sheet.ncols):
           labels.append(sheet.cell_value(0,col_index))
        return labels

    def loadandparsexls(self,path,tscol=None, datcol=None,maxnline=2**31):
        ## load excel file and parse contents

        #try:
           book=xlrd.open_workbook(path)
           sheets=book.sheet_names()
           choices=[]
           for sheet in sheets:
              choices.append(str(sheet))

           if not (tscol==None and datcol==None):
              sheet=book.sheet_by_index(0)
              tscolumn=tscol
              datacolumn=datcol
           else:
              # select sheet
              dlg = wx.SingleChoiceDialog(self,'Worksheets', 'Select worksheet', choices)
              if dlg.ShowModal() != wx.ID_OK:
                 return False,[],[],[]
              selectedsheet=dlg.GetSelection() # sheet index
              dlg.Destroy()
              sheet=book.sheet_by_index(selectedsheet)
   
              # select column
              selections=[]
              for col_index in range(0,sheet.ncols):
                 s=str(col_index)+':"'+str(sheet.cell_value(0,col_index))+'"' # include first line value in selections
                 selections.append(s)
              dlg = wx.SingleChoiceDialog(self,'Columns', 'Select timestamp column', selections)
              if dlg.ShowModal() != wx.ID_OK:
                 return False,[],[],[]
              tscolumn=dlg.GetSelection() # not string but index
              dlg.Destroy()
              dlg = wx.SingleChoiceDialog(self,'Columns', 'Select data column', selections)
              if dlg.ShowModal() != wx.ID_OK:
                 print('dlg.ShowModal()='+str(dlg.ShowModal()))
                 return False,[],[],[]
              datacolumn=dlg.GetSelection() # not string but index
              dlg.Destroy()
              dlg = wx.MessageDialog(parent=self,caption='Rows', message='Skip first row?',style=wx.YES|wx.NO)
              if dlg.ShowModal() == wx.ID_YES:
                 istart=1
              else: 
                 istart=0
   
              dlg.Destroy()
      
           # fill arrays
           datx=[]
           daty=[]
           datt=[]
           for row_index in range(istart,min(sheet.nrows,maxnline)):
              x=sheet.cell_value(row_index,tscolumn)
              y=sheet.cell_value(row_index,datacolumn)
              if not (x=='' or y==''):
                 try:
                    lx=float(x)
                    ly=float(y)
                    lt=float(lx)%1.0
                    datx.append(lx)
                    daty.append(ly)
                    datt.append(lt)
                 except:
                    #print 'skipping row '+str(row_index)+' (not parsable)'
                    pass
              else:
                 #print 'skipping row '+str(row_index)+' (empty)'
                 pass

           return True, datx, daty, datt

        #except:
           #return False,[],[],[]

    def loadandparsecsvheader(self,fname):
        f=open(fname,'r')
        lines=f.readlines()
        f.close()

        # determine seperator character based on first line
        hassemicolon=not lines[0].find(';')==-1
        hastab=not lines[0].find('\t')==-1
        hascomma=not lines[0].find(',')==-1
        hasspace=not lines[0].find(' ')==-1

        if hassemicolon:
           separator=';'
           cull1=','
           cull2='\t'
           cull3='\t'
        elif hascomma:
           separator=','
           cull1='\t'
           cull2=';'
           cull3='\t'
        elif hastab:
           separator='\t'
           cull1=','
           cull2=';'
           cull3=';'
        elif hasspace:
           separator=' '
           cull1='\t'
           cull2=','
           cull3=';'
        else:
           return False,[],[],[]

        # get column headers
        labels=lines[0].replace(cull1,'').replace(cull2,'').replace(cull3,'').replace('\r','').replace('\n','').split(separator)
        return labels

    def loadandparsecsv(self,fname,tscol=None, datcol=None,maxnline=2**31):
        f=open(fname,'r')
        lines=f.readlines()
        f.close()

        # determine seperator character based on first line
        hassemicolon=not lines[0].find(';')==-1
        hastab=not lines[0].find('\t')==-1
        hascomma=not lines[0].find(',')==-1
        hasspace=not lines[0].find(' ')==-1


        if hassemicolon:
           separator=';'
           cull1=','
           cull2='\t'
           cull3='\t'
        elif hascomma:
           separator=','
           cull1='\t'
           cull2=';'
           cull3='\t'
        elif hastab:
           separator='\t'
           cull1=','
           cull2=';'
           cull3=';'
        elif hasspace:
           separator=' '
           cull1='\t'
           cull2=','
           cull3=';'
        else:
           return False,[],[],[]

        # select column
        selectionstrings=lines[0].replace(cull1,'').replace(cull2,'').replace(cull3,'').replace('\r','').replace('\n','').split(separator)
 
        if len(selectionstrings)==2:
           # two columns -> first is timestamp and second is data\
           tscolumn=0
           datacolumn=1
           # assume first line to contain data; if not, it will be ignored by
           # the parser
           istart=0
        else:
           if not (tscol==None or datcol==None):
              tscolumn=tscol
              datacolumn=datcol
              istart=0 # if it is a label, it will be skipped anyway
           else:
              # more columns -> let user select 
              selections=[]
              for i in range(len(selectionstrings)):
                 selections.append(str(i)+':'+conversion.cleanupstring(selectionstrings[i]))
              dlg = wx.SingleChoiceDialog(self,'Columns', 'Select timestamp column', selections)
              if dlg.ShowModal() != wx.ID_OK:
                 return False,[],[],[]
              tscolumn=dlg.GetSelection() # not string but index
              dlg.Destroy()
              dlg = wx.SingleChoiceDialog(self,'Columns', 'Select data column', selections)
              if dlg.ShowModal() != wx.ID_OK:
                 print('dlg.ShowModal()='+str(dlg.ShowModal()))
                 return False,[],[],[]
              datacolumn=dlg.GetSelection() # not string but index
              dlg.Destroy()
              dlg = wx.MessageDialog(parent=self,caption='Rows', message='Skip first row?',style=wx.YES|wx.NO)
              if dlg.ShowModal() == wx.ID_YES:
                 istart=1
              else:
                 istart=0
              dlg.Destroy()

        # fill arrays
        datx=[]
        daty=[]
        datt=[]
        for row_index in range(istart,min(len(lines),maxnline)):
           ldat=lines[row_index].replace(cull1,'').replace(cull2,'').replace(cull3,'').replace('\r','').replace('\n','').split(separator)
           try:
              x=ldat[tscolumn]
              y=ldat[datacolumn]
           except:
              print('data error; ldat='+str(ldat))
              x=''; y=''
           if not (x=='' or y==''):
              try:
                 try:
                    lx=float(x)
                 except:
                    lx=conversion.string2exceldate(x)
                 ly=float(y)
                 lt=float(lx)%1.0
                 #print 'lx='+str(lx)+' ly='+str(ly)+' lt='+str(lt)
                 datx.append(lx)
                 daty.append(ly)
                 datt.append(lt)
              except:
                 #print 'skipping row '+str(row_index)+' (not parsable)   x="'+x+'" y="'+y+'"'
                 pass
           else:
              #print 'skipping row '+str(row_index)+' (empty)'
              pass

        return True, datx, daty, datt
 
    def parsefile(self,dat):
        datx=[]
        daty=[]
        datt=[]
        i=-1
        dlg = wx.ProgressDialog('Progress', 'Parsing data file', 2)
        if (dat[0].replace('\n','').replace('\r','')=='datetime;val'):
           # Exeter format
           #print 'Exeter format'
           dat.pop(0)
           for lline in dat:
              line=lline.replace('\r','').replace('\n','').split(';')
              #print "line="+lline
              val=float(line[1])
              daty.append(val)
              dats=line[0].split(' ')[0].split('-')
              tims=line[0].split(' ')[1].split(':')
              dt=datetime.datetime(int(dats[0]),int(dats[1]),int(dats[2]),int(tims[0]),int(tims[1]),int(tims[2]))                          
              datx.append(conversion.date2exceldate(dt))
              datt.append(float(tims[0])/24.0+float(tims[1])/1440.0+float(tims[2])/86400.0)
        else:
           # my format
           #print 'my format'
           irefday=conversion.getrefdat()

           shift=0.0
           for lline in dat:
              #print "lline="+str(lline)
              line=lline.replace('\r','').replace('\n','').replace(',',';').split(';')
              #print "line="+str(line)
              try:
                 day=float(line[0])
                 if (day<1000.0):
                    shift=irefday # prevent date conversion errors for files which have
                                  # weekday numbers instead of excel date numbers
                 val=float(line[1])
                 daty.append(val)
                 datx.append(float(line[0])+shift)
                 datt.append(float(line[0]))
              except:
                 print('ignoring line:"'+str(line)+'"')
        i+=1
        dlg.Update(i)
        dlg.Destroy()

        return datx, daty, datt

    def comparestats(self,evt):

        try:
           dt=float(self.dtfield.GetValue())/1440.0 # minutes -> days
        except:
           dt=5.0/1440.0
           dlg = wx.MessageDialog(self, 'Averaging window is invalid; using 5 minutes instead.', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()

        dlg = wx.ProgressDialog('Progress', 'Computing distributions', 6)
        dlg.Update(1)

        minx=conversion.string2exceldate(self.minxfield.GetValue())
        maxx=conversion.string2exceldate(self.maxxfield.GetValue())
        [lxdat,lydat,lidat]=mystat.subset(self.datx, self.daty, minx, maxx, -1e33, 1e33, 0, 1)

        #[lxdat,lydat]=mystat.subset(self.datx, self.daty, -1e33, 1e33, -1e33, 1e33, 0, 1)
        dlg.Update(2)
        lxdat2=mystat.dates2weekdays(lxdat) # get weekdays instead of dates
        dlg.Update(4)

        [cx, cy, cs] = mystat.getstats(lxdat2, lydat, dt)
        dlg.Update(5)


        lxdat3=mystat.dates2weekdays(self.datx2) # get weekdays instead of dates
        [sx, sy, ss] = mystat.getstats(lxdat3, self.daty2, dt)
        dlg.Update(5)

        [ttest_x, ttest_t, ttest_p, ttest_h]=mystat.getttest(lxdat2, lydat, lxdat3, self.daty2,dt)

        dlg.Destroy()

        # display results in window
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        self.comparisonwindow([[cx,cy,cs,t1],[sx,sy,ss,t2]],[ttest_x, ttest_p, ttest_h, 'two-sided T probability'])




    def comparisonwindow(self,sets,set2):
        self.cowinxlen=900
        self.cowinylen=630
        margin=4
        self.comparisonwin=wx.Frame(self,1, 'Flow pattern comparison', (150,150), (self.cowinxlen,self.cowinylen))
        self.comparisonwin.SetBackgroundColour((255,255,255))

        # panel
        x=margin; y=margin
        self.xgraphsize=self.cowinxlen-2*margin
        self.ygraphsize=self.cowinylen-2*margin-100-30-margin
        self.graphpanel3 = wx.Panel(self.comparisonwin, pos=(x,y), size=(self.xgraphsize,self.ygraphsize))
        self.panelredrawstatus["graphpanel3"]=True
        self.graphpanel3.Bind(wx.EVT_PAINT, lambda event: self.ongp3paint(self.comparisonwin,sets,set2,"graphpanel3"))

        # panel for T-test
        self.ttxgraphsize=self.xgraphsize
        self.ttygraphsize=100
        x=margin; y=2*margin+self.ygraphsize
        self.ttestpanel = wx.Panel(self.comparisonwin, pos=(x,y), size=(self.ttxgraphsize,self.ttygraphsize))
        self.panelredrawstatus["ttestpanel"]=True
        self.ttestpanel.Bind(wx.EVT_PAINT, lambda event: self.ongp3paint(self.comparisonwin,sets,set2,"ttestpanel"))

        xbs=80
        ybs=25
        x=margin
        y=self.cowinylen-margin-ybs
        self.showbghistcb=wx.CheckBox(self.comparisonwin,id=-1,label='Show background histograms',pos=(x,y))
        self.showbghistcb.SetValue(True)
        self.showbghistcb.Bind(wx.EVT_CHECKBOX, lambda event: self.ongp3paint(self.comparisonwin,sets,set2,"graphpanel3", forceredraw=True))

        x=self.cowinxlen-margin-xbs
        self.screenshotbutton2=wx.Button(self.comparisonwin, id=-1, label='Save image', pos=(x, y), size=(xbs,ybs))

        fname=''
        self.screenshotbutton2.Bind(wx.EVT_BUTTON, lambda event: screenshot.onTakeScreenShot(self,wx.EVT_BUTTON,self.comparisonwin,self.cowinxlen,self.cowinylen-ybs-margin,fname))

        self.comparisonwin.Show()


    def ongp3paint(self, win, sets, set2, panelname, forceredraw=False):

        if self.panelredrawstatus[panelname] or forceredraw:
           ldat=[]
           lcol=[]
           i=-1
           lsd=[]
           llab=[]
           minx=1e33; maxx=-1e33
           for set in sets:
              ldat.append([set[0],set[1]]) # x and y
              #print 'set x range='+str([min(set[0]),max(set[0])])
              i+=1
              lcol.append(i)
              lsd.append(set[2])
              llab.append(set[3])
              minx=min(minx,min(set[0]))
              maxx=max(maxx,max(set[0]))

           irefdat=conversion.getrefdat()

           try:
              dt=float(self.dtfield.GetValue())/1440.0 # minutes -> days
           except:
              dt=5.0/1440.0
              dlg = wx.MessageDialog(self, 'Averaging window is invalid; using 5 minutes instead.', 'Error', wx.ICON_ERROR)
              dlg.ShowModal()
   
           # min and max weekday
           d1=conversion.xldate_as_datetime(minx,0)
           d2=conversion.xldate_as_datetime(maxx,0)
           minwd=d1.weekday()
           maxwd=d2.weekday()+1
           if (maxx%1)<=0.51*dt:
              # last date is at stroke of midnight
              maxwd-=1

           self.panelbuffers[panelname+"cur"]=visual.drawcurves(self.graphpanel3,ldat,lcol,True,[irefdat+minwd, irefdat+maxwd],[-1,-1],self.xgraphsize,self.ygraphsize, False, [0,0], legend=llab, sd=lsd,fontsize=14,levels=True)

           self.panelbuffers[panelname+"hist"]=visual.drawhistcurves(self.ttestpanel,[set2[0],set2[1]],2,True,[irefdat+minwd, irefdat+maxwd],[-1,1],self.ttxgraphsize,self.ttygraphsize, set2[2], legend=set2[3], levels=True, drawbghist=self.showbghistcb.GetValue())
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(self.graphpanel3,self.panelbuffers[panelname+"cur"])
           visual.drawBitmap(self.ttestpanel,self.panelbuffers[panelname+"hist"])

    def computestats(self,evt):
   
        miny=float(self.minyfield.GetValue())
        maxy=float(self.maxyfield.GetValue())
        mint=conversion.timstr2num(self.mintfield.GetValue())
        maxt=conversion.timstr2num(self.maxtfield.GetValue())
        miny2=float(self.minyfield2.GetValue())
        maxy2=float(self.maxyfield2.GetValue())
        mint2=conversion.timstr2num(self.mintfield2.GetValue())
        maxt2=conversion.timstr2num(self.maxtfield2.GetValue())

        n=1000
        [self.data_Q, self.data_sigma, self.data_derQ, self.data_dersigma, mean, sd, self.data_epsilons] = mystat.gettrunccurves(self.datx, self.daty, miny, maxy, mint, maxt, n, 1,2)
        # compute corresponding synthetic curves
        [sx1,sy1,si1]=mystat.subset(self.datx, self.daty, self.minxzoom, self.maxxzoom, miny, maxy, mint, maxt)
        smean1=numpy.mean(sy1)
        sstd1=numpy.std(sy1)
        [self.syneps1,self.synE1,self.synS1,self.dersynE1, self.dersynS1]=mystat.getsynthcurves(n,smean1,sstd1)

        [self.data2_Q, self.data2_sigma, self.data2_derQ, self.data2_dersigma, data2_mean, data2_sd, self.data2_epsilons] = mystat.gettrunccurves(self.datx2, self.daty2, miny2, maxy2, mint2, maxt2, n, 2,2)
        # compute corresponding synthetic curves
        [sx2,sy2,si2]=mystat.subset(self.datx2, self.daty2, self.minxzoom2, self.maxxzoom2, miny2, maxy2, mint2, maxt2)
        smean2=numpy.mean(sy2)
        sstd2=numpy.std(sy2)
        [self.syneps2,self.synE2,self.synS2,self.dersynE2, self.dersynS2]=mystat.getsynthcurves(n,smean2,sstd2)

        self.showtrunccurvewin()

    def showtrunccurvewin(self):
        self.tcwinxlen=900
        self.tcwinylen=630
        margin=4
        self.trunccurvewin=wx.Frame(self,1, 'Truncation statistics', (150,150), (self.tcwinxlen,self.tcwinylen))
        self.trunccurvewin.SetBackgroundColour((255,255,255))

        self.gpxsize=self.tcwinxlen-2*margin
        self.gpysize=(self.tcwinylen-3*margin-30-30)/2

        # panels
        x=margin; y=margin
        self.graphpanel1 = wx.Panel(self.trunccurvewin, pos=(x,y), size=(self.gpxsize,self.gpysize))
        self.panelredrawstatus["graphpanel1"]=True
        self.graphpanel1.Bind(wx.EVT_PAINT, lambda event: self.ongp1paint(wx.EVT_PAINT,"graphpanel1"))
        y+=self.gpysize+2*margin
        self.graphpanel2 = wx.Panel(self.trunccurvewin, pos=(x,y), size=(self.gpxsize,self.gpysize))
        self.panelredrawstatus["graphpanel2"]=True
        self.graphpanel2.Bind(wx.EVT_PAINT, lambda event: self.ongp2paint(wx.EVT_PAINT,"graphpanel2"))

        xbs=80
        ybs=25

        x=240+margin
        y=self.tcwinylen-20-margin
        self.showsynth1=wx.CheckBox(self.trunccurvewin,id=-1,label='Show synth 1',pos=(x,y))
        self.showsynth1.SetValue(False)
        self.showsynth1.Bind(wx.EVT_CHECKBOX, lambda event: self.ongp12paint(wx.EVT_CHECKBOX,forceredraw=True))

        x+=140
        self.showsynth2=wx.CheckBox(self.trunccurvewin,id=-1,label='Show synth 2',pos=(x,y))
        self.showsynth2.SetValue(False)
        self.showsynth1.Bind(wx.EVT_CHECKBOX, lambda event: self.ongp12paint(wx.EVT_CHECKBOX,forceredraw=True))

        x=self.tcwinxlen-margin-xbs
        y=self.tcwinylen-margin-ybs
        self.screenshotbutton3=wx.Button(self.trunccurvewin, id=-1, label='Save image', pos=(x, y), size=(xbs,ybs))

        fname=''
        self.screenshotbutton3.Bind(wx.EVT_BUTTON, lambda event: screenshot.onTakeScreenShot(self,wx.EVT_BUTTON,self.trunccurvewin,self.tcwinxlen,self.tcwinylen-ybs-margin,fname))
 
        self.trunccurvewin.Show()

 

    def getdistpar(self,evt):
        # get parameters FROM THE OTHER PATTERN
        try:
           miny=float(self.minyfield2.GetValue())
           maxy=float(self.maxyfield2.GetValue())
           mint=conversion.timstr2num(self.mintfield2.GetValue())
           maxt=conversion.timstr2num(self.maxtfield2.GetValue())
           [sx,sy,si]=mystat.subset(self.datx2, self.daty2, self.minxzoom2, self.maxxzoom2, miny, maxy, mint, maxt)
   
           mean=numpy.mean(sy)
           sd=numpy.std(sy)
           trunc=min(sy)
   
           self.tndmeanfield.SetValue(str(mean))
           self.tndsdfield.SetValue(str(sd))
           self.truncfield.SetValue(str(trunc))
        except:
           dlg = wx.MessageDialog(self, 'Invalid value and/or time range entries', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()
        

    def getdistpar2(self,evt):
        # get parameters FROM THE OTHER PATTERN
        try:
           miny=float(self.minyfield.GetValue())
           maxy=float(self.maxyfield.GetValue())
           mint=conversion.timstr2num(self.mintfield.GetValue())
           maxt=conversion.timstr2num(self.maxtfield.GetValue())
           [sx,sy,si]=mystat.subset(self.datx, self.daty, self.minxzoom, self.maxxzoom, miny, maxy, mint, maxt)
   
           mean=numpy.mean(sy)
           sd=numpy.std(sy)
           trunc=min(sy)
   
           self.tndmeanfield2.SetValue(str(mean))
           self.tndsdfield2.SetValue(str(sd))
           self.truncfield2.SetValue(str(trunc))
        except:
           dlg = wx.MessageDialog(self, 'Invalid value and/or time range entries', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()

    def graphwindow(self,evt):
        xsize=self.xsize
        ysize=650
        self.graphwindow=wx.Frame(self,1, 'Graphs', (50,50), (xsize,ysize))
        self.graphwindow.SetBackgroundColour((255,255,255))

        margin1=4
        margin2=20

        xl=self.xframesize
        xl2=self.xframesize2
        xl3=self.xframesize3

        yl=100
        yl2=self.zoompanelysize
 
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        x=margin1; y=margin1
        wx.StaticText(self.graphwindow,-1,'a)',pos=(x,y+4))
        y+=margin2
        self.gwfppanel = wx.Panel(self.graphwindow, pos=(x, y), size=(xl,yl))
        self.panelredrawstatus["gwfppanel"]=True
        self.gwfppanel.Bind(wx.EVT_PAINT, lambda event: self.onfppaint(wx.EVT_PAINT,self.gwfppanel,"gwfppanel"))
        y+=yl+margin1 
        wx.StaticText(self.graphwindow,-1,'b) '+t1,pos=(x,y+4))
        y+=margin2
        self.gwzoompanel = wx.Panel(self.graphwindow, pos=(x,y), size=(xl2,yl2))
        self.panelredrawstatus["gwzoompanel"]=True
        self.gwzoompanel.Bind(wx.EVT_PAINT, lambda event: self.onzoompaint(wx.EVT_PAINT,self.gwzoompanel,"gwzoompanel"))
        x+=2*margin1+xl2
        y-=margin2
        wx.StaticText(self.graphwindow,-1,'c)',pos=(x,y+4))
        y+=margin2
        self.gwhistpanel = wx.Panel(self.graphwindow, pos=(x,y), size=(xl3,yl2))
        self.panelredrawstatus["gwhistpanel"]=True
        self.gwhistpanel.Bind(wx.EVT_PAINT, lambda event: self.onhistpaint(wx.EVT_PAINT,self.gwhistpanel,"gwhistpanel",equal=True))
        x=margin1; y+=margin2+yl2
        wx.StaticText(self.graphwindow,-1,'d)',pos=(x,y+4))
        y+=margin2
        self.gwfppanel2 = wx.Panel(self.graphwindow, pos=(x, y), size=(xl,yl))
        self.panelredrawstatus["gwfppanel2"]=True
        self.gwfppanel2.Bind(wx.EVT_PAINT, lambda event: self.onfppaint2(wx.EVT_PAINT,self.gwfppanel2,"gwfppanel2"))
        y+=margin1+yl
        wx.StaticText(self.graphwindow,-1,'e) '+t2,pos=(x,y+4))
        y+=margin2
        self.gwzoompanel2 = wx.Panel(self.graphwindow, pos=(x,y), size=(xl2,yl2))
        self.panelredrawstatus["gwzoompanel2"]=True
        self.gwzoompanel2.Bind(wx.EVT_PAINT, lambda event: self.onzoompaint2(wx.EVT_PAINT,self.gwzoompanel2,"gwzoompanel2"))
        x+=2*margin1+xl2
        y-=margin2
        wx.StaticText(self.graphwindow,-1,'f)',pos=(x,y+4))
        y+=margin2
        self.gwhistpanel2 = wx.Panel(self.graphwindow, pos=(x,y), size=(xl3,yl2))
        self.panelredrawstatus["gwhistpanel2"]=True
        self.gwhistpanel2.Bind(wx.EVT_PAINT, lambda event: self.onhistpaint2(wx.EVT_PAINT,self.gwhistpanel2,"gwhistpanel2",equal=True))

        xbs=80
        ybs=25
        x=xsize-margin1-xbs
        y=ysize-margin1-ybs
        self.screenshotbutton=wx.Button(self.graphwindow, id=-1, label='Save image', pos=(x, y), size=(xbs,ybs))
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        #fname='timeseries_'+t1+'-'+t2+'_'+str(datetime.datetime.now()).replace(" ","_")+'.png'
        fname=''
        self.screenshotbutton.Bind(wx.EVT_BUTTON, lambda event: screenshot.onTakeScreenShot(self,wx.EVT_BUTTON,self.graphwindow,xsize,ysize-ybs-margin1,fname))

        self.graphwindow.Show()

    def graphwindow2(self,evt):
        xsize=self.xsize
        ysize=650
        self.graphwindow2=wx.Frame(self,1, 'Graphs', (50,50), (xsize,ysize))
        self.graphwindow2.SetBackgroundColour((255,255,255))

        margin1=4
        margin2=20

        xl2=self.xframesize2
        xl3=self.xframesize3

        yl2=self.zoompanelysize
 
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        x=margin1; y=margin1
        wx.StaticText(self.graphwindow2,-1,'a) '+t1,pos=(x,y+4))
        y+=margin2
        self.gwzoompanel = wx.Panel(self.graphwindow2, pos=(x,y), size=(xl2,yl2))
        self.panelredrawstatus["gwzoompanel"]=True 
        self.gwzoompanel.Bind(wx.EVT_PAINT, lambda event: self.onzoompaint(wx.EVT_PAINT,self.gwzoompanel,"gwzoompanel"))
        x+=2*margin1+xl2
        y-=margin2
        wx.StaticText(self.graphwindow2,-1,'b)',pos=(x,y+4))
        y+=margin2
        self.gwhistpanel = wx.Panel(self.graphwindow2, pos=(x,y), size=(xl3,yl2))
        self.panelredrawstatus["gwhistpanel"]=True 
        self.gwhistpanel.Bind(wx.EVT_PAINT, lambda event: self.onhistpaint(wx.EVT_PAINT,self.gwhistpanel,"gwhistpanel",equal=True))
        x=margin1; y=margin1+yl2+margin2+margin1
        wx.StaticText(self.graphwindow2,-1,'c) '+t2,pos=(x,y+4))
        y+=margin2
        self.gwzoompanel2 = wx.Panel(self.graphwindow2, pos=(x,y), size=(xl2,yl2))
        self.panelredrawstatus["gwzoompanel2"]=True 
        self.gwzoompanel2.Bind(wx.EVT_PAINT, lambda event: self.onzoompaint2(wx.EVT_PAINT,self.gwzoompanel2,"gwzoompanel2"))
        x+=2*margin1+xl2
        y-=margin2
        wx.StaticText(self.graphwindow2,-1,'d)',pos=(x,y+4))
        y+=margin2
        self.gwhistpanel2 = wx.Panel(self.graphwindow2, pos=(x,y), size=(xl3,yl2))
        self.panelredrawstatus["gwhistpanel2"]=True 
        self.gwhistpanel2.Bind(wx.EVT_PAINT, lambda event: self.onhistpaint2(wx.EVT_PAINT,self.gwhistpanel2,"gwhistpanel2",equal=True))

        xbs=80
        ybs=25
        x=xsize-margin1-xbs
        y=ysize-margin1-ybs
        self.screenshotbutton=wx.Button(self.graphwindow2, id=-1, label='Save image', pos=(x, y), size=(xbs,ybs))
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        #fname='timeseries_red_'+t1+'-'+t2+'_'+str(datetime.datetime.now()).replace(" ","_")+'.png'
        fname=''
        self.screenshotbutton.Bind(wx.EVT_BUTTON, lambda event: screenshot.onTakeScreenShot(self,wx.EVT_BUTTON,self.graphwindow2,xsize,ysize-ybs-margin1,fname))

        self.graphwindow2.Show()

    def calcsinglejpp(self,win1,win2,dat1,dat2):
        [minxx1,maxx1,miny1,maxy1,mint1,maxt1]=win1
        [minxx2,maxx2,miny2,maxy2,mint2,maxt2]=win2
        [datx1,daty1]=dat1
        [datx2,daty2]=dat2

        [sx1,sy1,si1]=mystat.subset(datx1, daty1, minxx1,maxx1,miny1,maxy1,mint1,maxt1)
        sy1.sort()
        #print 'len sy1='+str(len(sy1))
        [sx2,sy2,si2]=mystat.subset(datx2, daty2, minxx2,maxx2,miny2,maxy2,mint2,maxt2)
        #print 'len sy2='+str(len(sy2))
        if len(sy1)<=1 or len(sy2)<=1:
           return False,[],[],[]
         
        sl=sorted(enumerate(sy2), key=operator.itemgetter(1))
        sy2=[]; newsi2=[]
        for item in sl:
           idx=item[0]
           val=item[1]
           sy2.append(val)
           newsi2.append(si2[idx])
        #print 'len sy2='+str(len(sy2))

        # always resample to length of set 2!
        if len(sy1)>0 and len(sy2)>0:
           jpp1dat=mystat.resample(sy1,len(sy2))
        else:
           jpp1dat=[]

        return True, jpp1dat,sy2,newsi2

    def calcsinglejppxx(self,win1,win2,dat1,dat2, timingcombofilter=False,fractionperc=100.0):
        [minxx1,maxx1,miny1,maxy1,mint1,maxt1]=win1
        [minxx2,maxx2,miny2,maxy2,mint2,maxt2]=win2
        [datx1,daty1]=dat1
        [datx2,daty2]=dat2

        [sx1,sy1,si1]=mystat.subset(datx1, daty1, minxx1,maxx1,miny1,maxy1,mint1,maxt1)
        sl=sorted(enumerate(sy1), key=operator.itemgetter(1))
        newsx1=[]
        for item in sl:
           idx=item[0]
           newsx1.append(sx1[idx]%1)  # no date; just time
           
        [sx2,sy2,si2]=mystat.subset(datx2, daty2, minxx2,maxx2,miny2,maxy2,mint2,maxt2)
        if len(sy1)<=1 or len(sy2)<=1:
           return False,[],[],[]

        sl=sorted(enumerate(sy2), key=operator.itemgetter(1))
        newsx2=[]
        for item in sl:
           idx=item[0]
           newsx2.append(sx2[idx]%1)  # no date; just time

        # always resample to length of set 2!
        resampledsx1=[]
        resampledsx2=[]
        ndat=len(sx1)
        n=len(sx2)
        #print 'ndat,n='+str([ndat,n])
        if ndat==0 or n==0:
           return [],[],[]
        elif ndat==n:
           for i in range(n):
              resampledsx1.append([newsx1[i]])
              resampledsx2.append([newsx2[i]])
        else:
           for i in range(n):
              j=i*float(ndat)/float(n)
              j1=int(math.floor(j)); j2=int(math.floor(j)+1)
               
              if j2>=ndat:
                 v=[newsx1[j1]]
              else:
                 v=[newsx1[j1],newsx1[j2]]
              resampledsx1.append(v)
              resampledsx2.append([newsx2[i]])
     
        # set filtering colors if required
        markercols=[]
        for i in range(len(resampledsx1)):
           if not timingcombofilter:
              markercols.append(0) # black 
           else:
              dx=(max(max(newsx1),max(newsx2))-min(min(newsx1),min(newsx2)))*fractionperc/100.0

              if abs(resampledsx1[i][0]-resampledsx2[i][0])<=dx:
                 #icol=1 # green 
              #elif len(resampledsx1[i])>1:
                 if len(resampledsx1[i])>1:
                    if abs(resampledsx1[i][1]-resampledsx2[i][0])<=dx :
                       icol=1 # green 
                    else:
                       icol=2 # red
                 else:
                    icol=1 # green
              else:
                 icol=2 # red
              markercols.append(icol)
    
        return True, resampledsx1,resampledsx2,markercols

    def calcjpp(self,evt):
        # file 1
        miny=float(self.minyfield.GetValue())
        maxy=float(self.maxyfield.GetValue())
        mint=conversion.timstr2num(self.mintfield.GetValue())
        maxt=conversion.timstr2num(self.maxtfield.GetValue())
        win1=[self.minxzoom, self.maxxzoom, miny, maxy, mint, maxt]
        dat1=[self.datx,self.daty]
        # file 2
        miny=float(self.minyfield2.GetValue())
        maxy=float(self.maxyfield2.GetValue())
        mint=conversion.timstr2num(self.mintfield2.GetValue())
        maxt=conversion.timstr2num(self.maxtfield2.GetValue())
        win2=[self.minxzoom2, self.maxxzoom2, miny, maxy, mint, maxt]
        dat2=[self.datx2,self.daty2]

        [result,jpp1dat,jpp2dat,jpp2i]=self.calcsinglejpp(win1,win2,dat1,dat2)
        if result:
           self.jpp1dat=jpp1dat
           self.jpp2dat=jpp2dat
           self.jpp2i=jpp2i
          
        self.createjppwindow(wx.EVT_PAINT)

    def calcnpp(self,evt):
        # compute data for normal probability plot and create window

        dlg = wx.ProgressDialog('Progress', 'Preparing normal probability plot', 2)
        dlg.Update(0)

        # file 1
        miny=float(self.minyfield.GetValue())
        maxy=float(self.maxyfield.GetValue())
        mint=conversion.timstr2num(self.mintfield.GetValue())
        maxt=conversion.timstr2num(self.maxtfield.GetValue())
        [sx1,sy1,si1]=mystat.subset(self.datx, self.daty, self.minxzoom, self.maxxzoom, miny, maxy, mint, maxt)
        self.npp1dat=mystat.normalprobabilityplotdata(sy1)
        dlg.Update(1)

        # file 2
        miny2=float(self.minyfield2.GetValue())
        maxy2=float(self.maxyfield2.GetValue())
        mint2=conversion.timstr2num(self.mintfield2.GetValue())
        maxt2=conversion.timstr2num(self.maxtfield2.GetValue())
        [sx2,sy2,si2]=mystat.subset(self.datx2, self.daty2, self.minxzoom2, self.maxxzoom2, miny2, maxy2, mint2, maxt2)
        self.npp2dat=mystat.normalprobabilityplotdata(sy2)
        dlg.Update(2)


        #print 'self.npp1dat='+str(self.npp1dat)
        #print 'self.npp2dat='+str(self.npp2dat)

        dlg.Destroy()

        self.createnppwindow(wx.EVT_PAINT)

    def getnnpfitcurve(self,x,y,istart, iend):
        [fit,residuals, rank, singular_values, rcond]=numpy.polyfit(x[istart:iend],y[istart:iend],1,full=True) # linear fit
        #print 'residuals='+str(residuals)
        #fit=numpy.polyfit(x[istart:iend],y[istart:iend],1) # linear fit

        self.nppfitslope=fit[0]
        self.nppfitintercept=fit[1]

        nused=iend-istart+1
        dx=(x[iend]-x[istart])/float(nused-1)
        xc=[]; yc=[]; sd=[]

        # calculation of standard deviation using:
        # http://en.wikipedia.org/wiki/Simple_linear_regression
        # with asymptotic assumption and q=1 because we just want to have the standard deviation
        xm=numpy.mean(x[istart:iend])
        ym=numpy.mean(y[istart:iend])
        sxxm=0.0; ssq=0.0; syym=0.0
        for i in range(istart,iend+1):
           sxxm+=(x[i]-xm)**2
           syym+=(y[i]-ym)**2
           ssq+=x[i]**2
        q=1
        for i in range(nused):
           lx=x[istart]+i*dx
           xc.append(lx)
           yc.append(fit[0]*lx+fit[1])
           if len(residuals)>0:
              sd.append(q*math.sqrt((residuals[0]/float(nused))*(1.0/float(nused)+(lx-xm)**2/sxxm)))
           else:
              sd.append(-1)

        # standard deviations of fit parameters:
        self.nppfitsdslope=math.sqrt((1.0/(float(nused)-2.0))*residuals[0]/sxxm)
 
        # different definitions; apparently the same results!
        self.nppfitsdintercept=math.sqrt((1.0/(float(nused)-2.0))*residuals[0])*math.sqrt((1.0/(float(nused)-2.0))+xm*xm/sxxm) # according to Wolfram Math World
        #self.nppfitsdintercept=math.sqrt((1.0/(float(nused)*(float(nused)-2.0)))*residuals[0]*ssq/sxxm) # according to wikipedia

        # coefficient of determination R^2
        # following http://en.wikipedia.org/wiki/Coefficient_of_determination
        self.nppfitR2=1.0-residuals[0]/syym
        #self.nppfitR2=1.0-residuals[0]/sxxm ! ERROR corrected 20120702 replacing sxxm with syym

        self.nppfitSSR=residuals[0]
        self.nppfitMSE=residuals[0]/float(nused)
        self.nppfitMSE_maxx=residuals[0]/float(nused)/max(x)
        self.nppfitRMSE=math.sqrt(residuals[0]/float(nused))
        self.nppfitRMSE_maxx=math.sqrt(residuals[0]/float(nused))/max(x)

        return [xc,yc],sd


    def getnnpfitcurvefrac(self,x,y,tx,ty,istart, iend,filteringtype,consistentfraction,ncluster=5,nclusterkeep=4,criterion='',subdivisionmethod='',timingcombofraction=1.0):
        # based on getnnpfitcurve
        # expansion: throw away points which deviate most from the best fit
        #            process in one percent steps
        #            retain best consistentfraction %

        # make local copies of arrays:
        localx=x[:]
        localy=y[:]
        ln=len(x)
        originalnodes=[]
        for i in range(ln):
           originalnodes.append(i)
 
        [fit,residuals, rank, singular_values, rcond]=numpy.polyfit(localx[istart:iend],localy[istart:iend],1,full=True) # linear fit
        slope=fit[0]
        intercept=fit[1]
        #print 'initial slope,intercept='+str([slope,intercept])

        markercolors=[]
        if filteringtype=='none':
           jstart=istart
           jend=iend
        elif filteringtype=='dev':
           # cull if required and resample:
           # first determine steps
           sizes=[]
           nused=iend-istart+1
           for istep in range(100,consistentfraction,-1):
              sizes.append(int(float(nused)*float(istep)/100.0))
           newsizes=list(set(sizes))
           newsizes.sort()
           newsizes.reverse()
           #print 'newsizes='+str(newsizes)
   
           # then do culling and refitting
           jstart=istart; jend=iend
           for istep in range(len(newsizes)):
              # cull
              newsize=newsizes[istep]
              differences=[]
              for i in range(len(localx)):
                  pred=intercept+slope*localx[i]
                  obs=localy[i]
                  differences.append(abs(obs-pred))
              sl=sorted(enumerate(differences), key=operator.itemgetter(1))
              llx=[]; lly=[]; llo=[]
              n=0
              markercolors=[]
              for item in sl:
                 idx=item[0]
                 if (idx>=jstart and idx<=jend):
                    llx.append(localx[idx])
                    lly.append(localy[idx])
                    llo.append(originalnodes[idx]) # remember original node number
                    markercolors.append(2)
                    n+=1
                 if n>newsize:
                    # do not include worst fitting points
                    break
              # update jstart and jend
              jstart=0
              jend=len(llx)-1
              localx=llx[:]
              localy=lly[:]
              originalnodes=llo[:]

              # recompute fit
              [fit,residuals, rank, singular_values, rcond]=numpy.polyfit(localx[jstart:jend],localy[jstart:jend],1,full=True) # linear fit
              slope=fit[0]
              intercept=fit[1]


        elif filteringtype=='clust':

           jstart=istart; jend=iend
           ln=len(localx)
           if ncluster==-1:
              ncluster=int(math.sqrt(ln))

           # determine cluster bounds
           dx=float(ln)/float(ncluster)
           if subdivisionmethod=='equal size':
              # distribute points equally over sections
              clusterbounds=[[0,int(dx)-1]] # start with first point
              for i in range(1,ncluster-1):
                 iprev=clusterbounds[-1][1]
                 clusterbounds.append([iprev+1,int((i+1)*dx-1)])
              clusterbounds.append([clusterbounds[-1][1]+1,ln-1]) # include last point
           elif subdivisionmethod=='optimal':
              # make optimal subdivision of coherent (i.e. close in (a,b)-space) sections
              clusterbounds=blockanalysis.determineclusters(self,localx,localy,ncluster,5,self.verbose)
           else:
              print('subdivisionmethod error')
              sys.exit(1)

           # get slope a and intercept b for each cluster
           clusta=[]; clustb=[]; clustres=[]
           for i in range(ncluster):
              kstart=clusterbounds[i][0]
              kend=clusterbounds[i][1]
              [cfit,cresiduals, crank, csingular_values, crcond]=numpy.polyfit(localx[kstart:kend],localy[kstart:kend],1,full=True) # linear fit
              clusta.append(cfit[0])
              clustb.append(cfit[1])
              clustres.append(cresiduals)
              #print 'cluster '+str(i)+' a,b='+str(cfit)

           if criterion=='linear fit':
              # a,b sets line up -> get fit and determine deviations for each cluster/section
              [lfit,lresiduals, lrank, lsingular_values, lrcond]=numpy.polyfit(clusta,clustb,1,full=True) # linear fit
              dev=[]
              for i in range(ncluster):
                 predb=lfit[0]*clusta[i]+lfit[1]
                 dev.append(abs(predb-clustb[i]))
           elif criterion=='median a':
              med=numpy.median(clusta)
              dev=[]
              for i in range(ncluster):
                 dev.append(abs(med-clusta[i]))
           else:
              print('criterion error')
              sys.exit(1)

           # sort for increasing deviation
           sl=sorted(enumerate(dev), key=operator.itemgetter(1))
           ln=0
           localx=[]; localy=[]; originalnodes=[]; markercolors=[]
           icols=[2,6] # marker colors, to visually distinguish the individual sections of used nodes
           for item in sl:
              idx=item[0]
              kstart=clusterbounds[idx][0]
              kend=clusterbounds[idx][1]
              localx+=x[kstart:kend+1]
              localy+=y[kstart:kend+1]
              originalnodes+=range(kstart,kend+1)
              m=idx%2
              markercolors+=[icols[m]]*(kend-kstart+1)
              ln+=1
              if ln>=nclusterkeep:
                 break
           jstart=0
           jend=len(localx)-1
               
           # recompute fit
           [fit,residuals, rank, singular_values, rcond]=numpy.polyfit(localx[jstart:jend],localy[jstart:jend],1,full=True) # linear fit
           slope=fit[0]
           intercept=fit[1]

        elif filteringtype=='time':
           # timing combo filter

           originalnodes=[]
           markercolors=[]
           # determine timing deviations

           # check if second set contains lists (pairs of node references) in case that
           # the sets are not of equal length
           listflag=len(tx[0])>1
           #print 'tx='+str(tx)
           #print 'ty='+str(ty)
           devfrac=(max(max(max(tx)),max(max(ty)))-min(min(min(tx)),min(min(ty))))*timingcombofraction/100.0
           #print 'devfrac='+str(devfrac)+' range='+str(min(min(min(tx)),min(min(ty))))+' - '+str(max(max(max(tx)),max(max(ty))))
           for i in range(len(tx)):
              if listflag:
                 if abs(tx[i][0]-ty[i][0])<=devfrac:
                    if len(tx[i])>1:
                       if abs(tx[i][1]-ty[i][0])<=devfrac:
                          # deviation is small enough -> accept
                          originalnodes.append(i)
                          localx.append(x[i])
                          localy.append(y[i])
                          originalnodes.append(i)
                          markercolors.append(2)
                    else:
                       # deviation is small enough -> accept
                       originalnodes.append(i)
                       localx.append(x[i])
                       localy.append(y[i])
                       originalnodes.append(i)
                       markercolors.append(2)
              else:
                 if abs(tx[i][0]-ty[i][0])<=devfrac:
                    # deviation is small enough -> accept
                    originalnodes.append(i)
                    localx.append(x[i])
                    localy.append(y[i])
                    originalnodes.append(i)
                    markercolors.append(2)
           jstart=0
           jend=len(localx)-1
           #print 'len='+str(len(localx))+'/'+str(len(x))
               
           # recompute fit
           [fit,residuals, rank, singular_values, rcond]=numpy.polyfit(localx[jstart:jend],localy[jstart:jend],1,full=True) # linear fit
           slope=fit[0]
           intercept=fit[1]
        else:
           print('getnnpfitcurvefrac: illegal filteringtype='+str(filteringtype))
           sys.exit(1)
        

        self.nppfitslope=fit[0]
        self.nppfitintercept=fit[1]

        nused=jend-jstart+1

        if filteringtype=='none' or filteringtype=='dev' or filteringtype=='clust' or filteringtype=='time':
           # calculation of standard deviation using:
           # http://en.wikipedia.org/wiki/Simple_linear_regression
           # with asymptotic assumption and q=1 because we just want to have the standard deviation
           xm=numpy.mean(localx[jstart:jend])
           sxxm=0.0; ssq=0.0
           for i in range(jstart,jend+1):
              sxxm+=(localx[i]-xm)**2
              ssq+=localx[i]**2
   
           # MAKE FIT LINE FOR ORIGINAL DATA SPAN
           dx=(x[iend]-x[istart])/float(len(x)-1)
           xc=[]; yc=[]; sd=[]
           q=1
           for i in range(len(x)):
              lx=x[istart]+i*dx
              xc.append(lx)
              yc.append(fit[0]*lx+fit[1])
              if len(residuals)>0:
                 sd.append(q*math.sqrt((residuals[0]/float(nused))*(1.0/float(nused)+(lx-xm)**2/sxxm)))
              else:
                 sd.append(-1)

        # standard deviations of fit parameters:
        if len(residuals)>0:
           self.nppfitsdslope=math.sqrt((1.0/(float(nused)-2.0))*residuals[0]/sxxm)
 
           # different definitions; apparently the same results!
           self.nppfitsdintercept=math.sqrt((1.0/(float(nused)-2.0))*residuals[0])*math.sqrt((1.0/(float(nused)-2.0))+xm*xm/sxxm) # according to Wolfram Math World
           #self.nppfitsdintercept=math.sqrt((1.0/(float(nused)*(float(nused)-2.0)))*residuals[0]*ssq/sxxm) # according to wikipedia

           # coefficient of determination R^2
           # following http://en.wikipedia.org/wiki/Coefficient_of_determination
           self.nppfitR2=1.0-residuals[0]/sxxm
           self.nppfitSSR=residuals[0]
           try:
              self.nppfitMSE=residuals[0]/float(nused)
           except ZeroDivisionError:
              self.nppfitMSE=-999
           try:
              self.nppfitMSE_maxx=residuals[0]/float(nused)/max(localx)
           except ZeroDivisionError:
              self.nppfitMSE_maxx=-999
           try:
              self.nppfitRMSE=math.sqrt(residuals[0]/float(nused))
           except ZeroDivisionError:
              self.nppfitRMSE=-999
           try:
              self.nppfitRMSE_maxx=math.sqrt(residuals[0]/float(nused))/max(localx)
           except ZeroDivisionError:
              self.nppfitRMSE_maxx=-999
        else:
           self.nppfitsdslope=-1
           self.nppfitsdintercept=-1
           self.nppfitR2=-999
           self.nppfitSSR=-999
           self.nppfitMSE=-999
           self.nppfitMSE=-999
           self.nppfitRMSE=-999
           self.nppfitRMSE_maxx=-999

        # get complement of originalnodes: ignorednodes
        ignorednodes=[]
        for i in range(istart,iend+1):
           try:
              idx=originalnodes.index(i)
           except:
              ignorednodes.append(i)
        #print 'ignorednodes='+str(ignorednodes)

        return [xc,yc],sd, originalnodes,ignorednodes,markercolors

    def createjppwindow(self,evt):
        xsize=int(self.ysize-250)
        ysize=int(self.ysize-50)
      
        self.jppwindow=wx.Frame(self,-1, 'CFPD Plot', (50,50), (xsize,ysize))
        self.jppwindow.SetBackgroundColour((255,255,255))

        margin1=4
        margin2=20
        xbs=80
        ybs=25

        xl=int((xsize-2*margin1-xbs))
        yl=xl

        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
 
        x=margin1; y=margin1
        wx.StaticText(self.jppwindow,-1,'hori: '+t1+'   vert: '+t2,pos=(x,y+4))
        y+=margin2
        self.jpppanel = wx.Panel(self.jppwindow, pos=(x, y), size=(xl,yl))
        self.panelredrawstatus["jpppanel"]=True 
        self.jpppanelstaticbitmap = wx.StaticBitmap(self.jpppanel,-1,pos=(0,0), size=(xl,yl))


        # scroll bars for fit range
        y=int(ysize-margin1-6*ybs)
        x=margin1
        sbl=int((xsize-3*margin1-4*xbs)/2)
        sbh=12
        self.jnpfitrangesb11=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
                                        # position, thumbSize, range, pageSize
        self.jnpfitrangesb11.SetScrollbar(0, 1, len(self.jpp1dat), 1, refresh=True)
        self.jnpfitrangesb11.SetToolTip(wx.ToolTip("Min volume flow for set 1 for fit"))
        x+=sbl+margin1
        self.jnpfitrangesb12=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
        self.jnpfitrangesb12.SetScrollbar(len(self.jpp1dat)-1, 1, len(self.jpp1dat), 1, refresh=True)
        self.jnpfitrangesb12.SetToolTip(wx.ToolTip("Max volume flow for set 1 for fit"))

        x+=sbl+margin1
        self.jnpcheckfit=wx.CheckBox(self.jppwindow,id=-1,label='Fit',pos=(x,y))
        self.jnpcheckfit.SetValue(True)
        self.jnpcheckfit.Bind(wx.EVT_CHECKBOX, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))

        x+=int(0.5*xbs+margin1)
        self.jnpcheckmarker=wx.CheckBox(self.jppwindow,id=-1,label='Markers',pos=(x,y))
        self.jnpcheckmarker.SetValue(False)
        self.jnpcheckmarker.Bind(wx.EVT_CHECKBOX, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))

        self.jnpfitrangesb11.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpfitrangesb12.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))


        x=int(xsize-2*margin1-2*xbs)
        #y=ysize-margin1-ybs
        self.jppfitinfobutton=wx.Button(self.jppwindow, id=-1, label='Info', pos=(x, y), size=(int(0.5*xbs),ybs))
        self.jppfitinfobutton.Bind(wx.EVT_BUTTON,self.jppinfo)
        self.jppfitinfobutton.SetToolTip(wx.ToolTip("Display numerical information on data and fit."))
        x+=int(margin1+0.5*xbs)
        self.jppscreenshotbutton=wx.Button(self.jppwindow, id=-1, label='SaveI', pos=(x, y), size=(int(xbs/2),ybs))
        self.jppscreenshotbutton.Bind(wx.EVT_BUTTON, lambda event: self.savejppscreenshotandfitinfo(wx.EVT_BUTTON,self.jppwindow,xsize,yl+2*margin1+margin2))
        self.jppscreenshotbutton.SetToolTip(wx.ToolTip("Save image."))
        x+=int(margin1+0.5*xbs)
        self.jppsavedatabutton=wx.Button(self.jppwindow, id=-1, label='SaveD', pos=(x, y), size=(int(xbs/2),ybs))
        self.jppsavedatabutton.Bind(wx.EVT_BUTTON, self.savejppdata)
        self.jppsavedatabutton.SetToolTip(wx.ToolTip("Save data."))


        # scroll bars for zoom range
        y+=sbh+margin1
        x=margin1
        sbh=12
        self.jnpxzoomrangesb11=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
                                        # position, thumbSize, range, pageSize
        self.jnpxzoomrangesb11.SetScrollbar(0, 1, len(self.jpp1dat), 1, refresh=True)
        self.jnpxzoomrangesb11.SetToolTip(wx.ToolTip("Min volume flow for set 1 for zoom"))
        x+=sbl+margin1
        self.jnpxzoomrangesb12=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
        self.jnpxzoomrangesb12.SetScrollbar(len(self.jpp1dat)-1, 1, len(self.jpp1dat), 1, refresh=True)
        self.jnpxzoomrangesb12.SetToolTip(wx.ToolTip("Max volume flow for set 1 for zoom"))

        y+=sbh+margin1
        x=margin1
        self.jnpyzoomrangesb11=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
                                        # position, thumbSize, range, pageSize
        self.jnpyzoomrangesb11.SetScrollbar(0, 1, len(self.jpp2dat), 1, refresh=True)
        self.jnpyzoomrangesb11.SetToolTip(wx.ToolTip("Min volume flow for set 2 for zoom"))
        x+=sbl+margin1
        self.jnpyzoomrangesb12=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
        self.jnpyzoomrangesb12.SetScrollbar(len(self.jpp2dat)-1, 1, len(self.jpp2dat), 1, refresh=True)
        self.jnpyzoomrangesb12.SetToolTip(wx.ToolTip("Max volume flow for set 2 for zoom"))

        y+=sbh+margin1
        x=margin1
        self.jnptzoomrangesb11=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
                                        # position, thumbSize, range, pageSize
        self.jnptzoomrangesb11.SetScrollbar(0, 1,24,1, refresh=True)
        self.jnptzoomrangesb11.SetToolTip(wx.ToolTip("Min time for zoom"))
        x+=sbl+margin1
        self.jnptzoomrangesb12=wx.ScrollBar(self.jppwindow,-1,pos=(x,y),size=(sbl,sbh))
        self.jnptzoomrangesb12.SetScrollbar(24,1, 24,1, refresh=True)
        self.jnptzoomrangesb12.SetToolTip(wx.ToolTip("Max time for zoom"))

        self.jnpxzoomrangesb11.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpxzoomrangesb12.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpyzoomrangesb11.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpyzoomrangesb12.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnptzoomrangesb11.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnptzoomrangesb12.Bind(wx.EVT_SCROLL, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))

        x+=sbl+margin1
        y-=sbh+margin1
        self.jnpcheckincdeclines=wx.CheckBox(self.jppwindow,id=-1,label='Dev.l.',pos=(x,y))
        self.jnpcheckincdeclines.SetValue(False)
        self.jnpcheckincdeclines.Bind(wx.EVT_CHECKBOX, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpcheckincdeclines.SetToolTip(wx.ToolTip("Show absolute deviation lines with respect to ideal curve"))

        x+=int(0.6*xbs+margin1)
        self.jnpcheckperclines=wx.CheckBox(self.jppwindow,id=-1,label='Perc.l.',pos=(x,y))
        self.jnpcheckperclines.SetValue(False)
        self.jnpcheckperclines.Bind(wx.EVT_CHECKBOX, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpcheckperclines.SetToolTip(wx.ToolTip("Show percentage deviation lines relative to ideal curve"))

        x+=int(0.6*xbs+margin1)
        self.jnpcheckzoom=wx.CheckBox(self.jppwindow,id=-1,label='Zoom',pos=(x,y))
        self.jnpcheckzoom.SetValue(False)
        self.jnpcheckzoom.Bind(wx.EVT_CHECKBOX, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpcheckzoom.SetToolTip(wx.ToolTip("Select part of the CFPD diagram using the sliders."))


        x+=int(0.6*xbs+margin1)
        self.jppzoombutton=wx.Button(self.jppwindow, id=-1, label='Zoom', pos=(x, y), size=(int(0.5*xbs),ybs))
        self.jppzoombutton.Bind(wx.EVT_BUTTON,self.jppzoomwindow)
        self.jppzoombutton.SetToolTip(wx.ToolTip("Show time series for selected part of CFPD diagram."))


        x=int(2*sbl+3*margin1)
        y+=int(sbh+3*margin1)
        self.jnpfontsize=wx.TextCtrl(self.jppwindow,pos=(x,y),size=(int(0.35*xbs),ybs))
        self.jnpfontsize.SetValue('11')
        self.jnpfontsize.Bind(wx.EVT_TEXT, lambda event: self.onjpppaint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.jnpfontsize.SetToolTip(wx.ToolTip("Font size for grid labels"))
        x+=int(0.4*xbs+margin1)
        wx.StaticText(self.jppwindow,-1,'Font size',pos=(x,y+4))

        x+=int(0.6*xbs+margin1)
        self.jppplusbutton=wx.Button(self.jppwindow, id=-1, label='+', pos=(x, y), size=(int(0.3*xbs),ybs))
        self.jppplusbutton.Bind(wx.EVT_BUTTON, lambda event: self.onjppbutton(wx.EVT_BUTTON,'plus',xl,yl))
        self.jppplusbutton.SetToolTip(wx.ToolTip("Increase number of shift/scale lines."))

        x+=int(0.3*xbs+margin1)
        self.jppminusbutton=wx.Button(self.jppwindow, id=-1, label='-', pos=(x, y), size=(int(0.3*xbs),ybs))
        self.jppminusbutton.Bind(wx.EVT_BUTTON, lambda event: self.onjppbutton(wx.EVT_BUTTON,'minus',xl,yl))
        self.jppminusbutton.SetToolTip(wx.ToolTip("Increase number of shift/scale lines."))

        x+=int(0.3*xbs+margin1)
        wx.StaticText(self.jppwindow,-1,'#Dev.l.',pos=(x,y+4))

        self.devlinefactor=1.0
        
        self.onjpppaint(wx.EVT_PAINT,xl,yl)


        self.jppwindow.Show()

    def onjppbutton(self,evt,op,xl,yl):

        if op=='plus':
           self.devlinefactor+=1
        else:
           self.devlinefactor-=1
         

        self.panelredrawstatus["jpppanel"]=True
        self.onjpppaint(wx.EVT_PAINT,xl,yl)

    def createnppwindow(self,evt):
        xsize=self.xsize
        ysize=self.ysize
        self.nppwindow=wx.Frame(self,-1, 'Normal Probability Plots', (50,50), (xsize,ysize))
        self.nppwindow.SetBackgroundColour((255,255,255))

        margin1=4
        margin2=20
        xbs=80
        ybs=25

        xl=int((xsize-3*margin1)/2)
        yl=int(ysize-3*margin1-ybs-margin2)
 
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        x=margin1; y=margin1
        wx.StaticText(self.nppwindow,-1,'a) '+t1,pos=(x,y+4))
        y+=margin2
        self.npp1panel = wx.Panel(self.nppwindow, pos=(x, y), size=(xl,yl))
        self.panelredrawstatus["npp1panel"]=True 
        x+=xl+margin1; y=margin1 
        wx.StaticText(self.nppwindow,-1,'b) '+t2,pos=(x,y+4))
        y+=margin2
        self.npp2panel = wx.Panel(self.nppwindow, pos=(x,y), size=(xl,yl))
        self.panelredrawstatus["npp2panel"]=True 

        # scroll bars for fit range
        y=ysize-margin1-ybs
        x=margin1
        sbl=(xsize-5*margin1-2*xbs)/4
        sbh=12
        self.nnpfitrangesb11=wx.ScrollBar(self.nppwindow,-1,pos=(x,y),size=(sbl,sbh))
                                        # position, thumbSize, range, pageSize
        self.nnpfitrangesb11.SetScrollbar(0, 1, len(self.npp1dat[0]), 1, refresh=True)
        x+=sbl+margin1
        self.nnpfitrangesb12=wx.ScrollBar(self.nppwindow,-1,pos=(x,y),size=(sbl,sbh))
        self.nnpfitrangesb12.SetScrollbar(len(self.npp1dat[0])-1, 1, len(self.npp1dat[0]), 1, refresh=True)

        x+=sbl+margin1
        self.nnpcheckfit=wx.CheckBox(self.nppwindow,id=-1,label='Fit',pos=(x,y))
        self.nnpcheckfit.SetValue(True)
        self.nnpcheckfit.Bind(wx.EVT_CHECKBOX, lambda event: self.onnpp12paint(wx.EVT_PAINT,xl,yl,forceredraw=True))

        x+=30+margin1
        self.nnpchecksamescale=wx.CheckBox(self.nppwindow,id=-1,label='Scal',pos=(x,y))
        self.nnpchecksamescale.SetValue(True)
        self.nnpchecksamescale.Bind(wx.EVT_CHECKBOX, lambda event: self.onnpp12paint(wx.EVT_PAINT,xl,yl,forceredraw=True))


        x=xl+2*margin1
        self.nnpfitrangesb21=wx.ScrollBar(self.nppwindow,-1,pos=(x,y),size=(sbl,sbh))
        self.nnpfitrangesb21.SetScrollbar(0, 1, len(self.npp2dat[0]), 1, refresh=True)
        x+=sbl+margin1
        self.nnpfitrangesb22=wx.ScrollBar(self.nppwindow,-1,pos=(x,y),size=(sbl,sbh))
        self.nnpfitrangesb22.SetScrollbar(len(self.npp2dat[0])-1, 1, len(self.npp2dat[0]), 1, refresh=True)

        self.nnpfitrangesb11.Bind(wx.EVT_SCROLL, lambda event: self.onnpp1paint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.nnpfitrangesb12.Bind(wx.EVT_SCROLL, lambda event: self.onnpp1paint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.nnpfitrangesb21.Bind(wx.EVT_SCROLL, lambda event: self.onnpp2paint(wx.EVT_PAINT,xl,yl,forceredraw=True))
        self.nnpfitrangesb22.Bind(wx.EVT_SCROLL, lambda event: self.onnpp2paint(wx.EVT_PAINT,xl,yl,forceredraw=True))

        self.npp1panel.Bind(wx.EVT_PAINT, lambda event: self.onnpp1paint(wx.EVT_PAINT,xl,yl))
        self.npp2panel.Bind(wx.EVT_PAINT, lambda event: self.onnpp2paint(wx.EVT_PAINT,xl,yl))

        x=xsize-margin1-xbs
        y=ysize-margin1-ybs
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        d=str(datetime.datetime.now()).replace(" ","_")
        fname='normal_probability_plots_'+t1+'-'+t2+'_'+d+'.png'
        self.nppscreenshotbutton=wx.Button(self.nppwindow, id=-1, label='Save image', pos=(x, y), size=(xbs,ybs))
        fname=''
        self.nppscreenshotbutton.Bind(wx.EVT_BUTTON, lambda event: screenshot.onTakeScreenShot(self,wx.EVT_BUTTON,self.nppwindow,xsize,ysize-ybs-margin1,fname))


        self.nppwindow.Show()

    def savejppscreenshotandfitinfo(self,evt,window,xsize,ysize):
        t1=self.title1.GetValue()
        t2=self.title2.GetValue()
        d=str(datetime.datetime.now()).replace(" ","_")
        #fname='joint_probability_plot_'+t1+'-'+t2+'_'+d+'.png'
        fname=''
        screenshot.onTakeScreenShot(self,evt,window,xsize,ysize,fname)

        fname='joint_probability_plot_fit_'+t1+'-'+t2+'_'+d+'.txt'
        [slope,intercept,sdslope,sdintercept,perc,ndat]=self.getjppinfo()
        f=open(fname,'w')
        f.write(d+'\n')
        f.write('---------------------\n')
        f.write('set 1: '+t1+'\n')
        f.write('set 2: '+t2+'\n')
        f.write('slope: '+str(slope)+' sd='+str(sdslope)+'\n')
        f.write('y intercept: '+str(intercept)+' sd='+str(sdintercept)+'\n')
        f.write('n='+str(ndat)+'\n')
        f.write('coefficient of determination R^2='+str(self.nppfitR2)+'\n')
        f.write('percentage of data used: '+str(perc)+'%\n')
        f.write('SSR='+str(self.nppfitSSR))
        f.write('MSE='+str(self.nppfitMSE))
        f.write('MSE/x_max='+str(self.nppfitMSE_maxx))
        f.write('RMSE='+str(self.nppfitRMSE))
        f.write('RMSE/x_max='+str(self.nppfitRMSE_maxx))
        f.close()

    def getjppinfo(self):
        istart=self.jnpfitrangesb11.GetThumbPosition()
        iend=self.jnpfitrangesb12.GetThumbPosition()
        ndat=self.jnpfitrangesb11.GetRange()
        percdatause=100.0*float(iend-istart+1)/float(ndat)

        return self.nppfitslope,self.nppfitintercept,self.nppfitsdslope, self.nppfitsdintercept,percdatause, ndat

    def jppinfo(self,evt):

        [slope,intercept,sdslope,sdintercept,perc,ndat]=self.getjppinfo()

        lstr='Linear fit:\nslope='+str(slope)+'\n'
        lstr+='sd='+str(sdslope)+'\n'
        lstr+='y-intercept='+str(intercept)+'\n'
        lstr+='sd='+str(sdintercept)+'\n'
        lstr+='n='+str(ndat)+'\n'
        lstr+='R^2='+str(self.nppfitR2)+'\n'
        lstr+='data use='+str(perc)+'%'+'\n'
        lstr+='SSR='+str(self.nppfitSSR)+'\n'
        lstr+='MSE='+str(self.nppfitMSE)+'\n'
        lstr+='MSE/x_max='+str(self.nppfitMSE_maxx)+'\n'
        lstr+='RMSE='+str(self.nppfitRMSE)+'\n'
        lstr+='RMSE/x_max='+str(self.nppfitRMSE_maxx)+'\n'

        dlg = wx.MessageDialog(self.jppwindow, lstr, 'Info', wx.ICON_INFORMATION)
        dlg.ShowModal()

    def savejppdata(self,evt):
        #path = wx.SaveFileSelector('CFPD data file', '.csv', '')
        dlg = wx.FileDialog(self,message='Save CFPD data file', defaultDir=self.cwd, wildcard='*.csv', style=wx.FD_SAVE)
        if dlg.ShowModal() == wx.ID_OK:
           path=dlg.GetPath()
           self.cwd=fsconfig.strippath(path)
           print('self.cwd->"'+self.cwd+'"')
           fsconfig.saveconfig([self.cwd,self.appsize])
        else:
           path=''
        dlg.Destroy()
        if not (path==''):
           # save data
           f=open(path,'w')
           for i in range(len(self.jpp1dat)):
              f.write(str(self.jpp1dat[i])+','+str(self.jpp2dat[i])+'\n')
           f.close()

           # save fit
           istart=self.jnpfitrangesb11.GetThumbPosition()
           iend=self.jnpfitrangesb12.GetThumbPosition()
           [fitline,sd]=self.getnnpfitcurve(self.jpp1dat,self.jpp2dat,istart, iend)

           fitpath=path.replace('.csv','_fitline.csv')
           f=open(fitpath,'w')
           for i in range(len(self.jpp1dat)):
              f.write(str(fitline[0][i])+','+str(fitline[1][i])+'\n')
           f.close()

           # save parameters
           [slope,intercept,sdslope,sdintercept,perc,ndat]=self.getjppinfo()
           datpath=path.replace('.csv','_param.csv')
           lstr='Linear fit:\nslope='+str(slope)+'\n'
           lstr+='sd='+str(sdslope)+'\n'
           lstr+='y-intercept='+str(intercept)+'\n'
           lstr+='sd='+str(sdintercept)+'\n'
           lstr+='n='+str(ndat)+'\n'
           lstr+='R^2='+str(self.nppfitR2)+'\n'
           lstr+='data use='+str(perc)+'%'+'\n'
           lstr+='SSR='+str(self.nppfitSSR)+'\n'
           lstr+='MSE='+str(self.nppfitMSE)+'\n'
           lstr+='MSE/x_max='+str(self.nppfitMSE_maxx)+'\n'
           lstr+='RMSE='+str(self.nppfitRMSE)+'\n'
           lstr+='RMSE/x_max='+str(self.nppfitRMSE_maxx)+'\n'
           f=open(datpath,'w')
           f.write(lstr)
           f.close()

           # save label
           labpath=path.replace('.csv','_label.csv')
           lstr='a=%7.5f [-]\n' % slope
           lstr+='b=%7.5f [m3/hr]\n' % intercept
           f=open(labpath,'w')
           f.write(lstr)
           f.close()
            

    def onjpppaint(self, evt, xsize, ysize, forceredraw=False):

        if self.panelredrawstatus["jpppanel"] or forceredraw:
           minxy=min(min(self.jpp1dat),min(self.jpp2dat))
           maxxy=max(max(self.jpp1dat),max(self.jpp2dat))
   
           dxy=maxxy-minxy
           minxy-=0.05*dxy; maxxy+=0.05*dxy
   
           dat=[[[minxy,maxxy],[minxy,maxxy]]] # equal probability line
           markers=[]
           curvemarkers=[False] # no markers for equal prob line
           lsd=[[]]
   
           if (self.jnpcheckfit.GetValue()):
              # a fit should be made
              istart=self.jnpfitrangesb11.GetThumbPosition()
              iend=self.jnpfitrangesb12.GetThumbPosition()
              if (iend>istart):
                 [fitline,sd]=self.getnnpfitcurve(self.jpp1dat,self.jpp2dat,istart, iend)
                 dat.append(fitline)
                 lsd.append(sd)
                 markers=[[self.jpp1dat[istart],self.jpp2dat[istart]],[self.jpp1dat[iend],self.jpp2dat[iend]]]
                 curvemarkers.append(False) # no curve markers for fit line
                 ifitcurve=[0,0]
                 col=[1,2,0]
           else:
              ifitcurve=[0,0]
              col=[1,0]
   
           dat.append([self.jpp1dat,self.jpp2dat])
           curvemarkers.append(self.jnpcheckmarker.GetValue())
           lsd.append([])
   
           #print 'dat='+str(dat)
   
           if (self.jnpcheckzoom.GetValue()):
              # show zoom bands
              xstart=self.jpp1dat[self.jnpxzoomrangesb11.GetThumbPosition()]
              xend=self.jpp1dat[self.jnpxzoomrangesb12.GetThumbPosition()]
              ystart=self.jpp2dat[self.jnpyzoomrangesb11.GetThumbPosition()]
              yend=self.jpp2dat[self.jnpyzoomrangesb12.GetThumbPosition()]
              if (xstart>xend):
                 xstart=xend
              if (ystart>yend):
                 ystart=yend
              zoombands=[xstart,xend,ystart,yend]
              #print 'zoombands='+str(zoombands)
           else:
              zoombands=[]
   
           if not self.jnpcheckincdeclines.GetValue():
              # do not show percentage deviation lines
              ifitcurve[0]=-1
           if not self.jnpcheckperclines.GetValue():
              # do not show percentage deviation lines
              ifitcurve[1]=-1
   
           try:
              fontsize=int(self.jnpfontsize.GetValue())
           except:
              fontsize=11
   
           self.panelbuffers["jpppanel"]=visual.drawcurves(self.jpppanelstaticbitmap,dat,col,False,[minxy,maxxy],[minxy,maxxy],xsize,ysize, False, [0,0], sd=lsd,legend=[], fontsize=fontsize, levels=True, vergrid=True, checkskip=False, linethickness=2, markers=markers,markercols=[1,1],markersize=2,curvemarkers=curvemarkers, zoombands=zoombands, ifitcurve=ifitcurve,labelgridline=True,devlinefactor=self.devlinefactor, horvergridsame=True) 
           self.panelredrawstatus["jpppanel"]=False
        else:
           visual.drawBitmap(self.jpppanel,self.panelbuffers["jpppanel"])


        self.jpppanelstaticbitmap.SetBitmap(self.panelbuffers["jpppanel"])
           

    def jppzoomwindow(self,evt):
        xsize=self.xsize
        ysize=650
        self.jppzoomwindow=wx.Frame(self,1, 'Zoom', (50,50), (xsize,ysize))
        self.jppzoomwindow.SetBackgroundColour((255,255,255))

        margin1=4

        xl=self.xframesize

        yl=300
        self.yframesizejpp=yl

        x=margin1
        y=margin1
 
        self.jppzoompanel = wx.Panel(self.jppzoomwindow, pos=(x,y), size=(xl,yl))
        self.panelredrawstatus["jppzoompanel"]=True 
        self.jppzoompanel.Bind(wx.EVT_PAINT, lambda event: self.onjppzoompaint(wx.EVT_PAINT,self.jppzoompanel))

        xbs=80
        ybs=25
        x=xsize-margin1-xbs
        y=ysize-margin1-ybs
        self.screenshotbutton=wx.Button(self.jppzoomwindow, id=-1, label='Save image', pos=(x, y), size=(xbs,ybs))

        fname=''
        self.screenshotbutton.Bind(wx.EVT_BUTTON, lambda event: screenshot.onTakeScreenShot(self,wx.EVT_BUTTON,self.jppzoomwindow,xsize,yl+2*margin1,fname))

        self.jppzoomwindow.Show()

    def onjppzoompaint(self,evt,panel):

        # generate markers for zoomed region
        xstart=self.jpp1dat[self.jnpxzoomrangesb11.GetThumbPosition()]
        xend=self.jpp1dat[self.jnpxzoomrangesb12.GetThumbPosition()]
        ystart=self.jpp2dat[self.jnpyzoomrangesb11.GetThumbPosition()]
        yend=self.jpp2dat[self.jnpyzoomrangesb12.GetThumbPosition()]
        tstart=float(self.jnptzoomrangesb11.GetThumbPosition())/24.0
        tend=float(self.jnptzoomrangesb12.GetThumbPosition())/24.0
        [sx,sy,si]=mystat.subset(self.jpp1dat, self.jpp2dat,xstart,xend, ystart,yend,tstart,tend)

        miny=float(self.minyfield2.GetValue())
        maxy=float(self.maxyfield2.GetValue())
        mint=conversion.timstr2num(self.mintfield2.GetValue())
        maxt=conversion.timstr2num(self.maxtfield2.GetValue())
        [sx2,sy2,si2]=mystat.subset(self.datx2, self.daty2, self.minxzoom2, self.maxxzoom2, miny, maxy, mint, maxt)

     
        markers=[]; markercols=[]
        for i in range(len(sy)):
           indx=si[i]
           jndx=self.jpp2i[indx]
           #print 'i,indx,jndx='+str([i,indx,jndx])
           x=self.datx2[jndx]
           y=self.daty2[jndx]
           markers.append([x,y])
           markercols.append(1)
 
        #if len(markers)<100:
           #print 'markers='+str(markers)

        visual.drawcurves(panel,[[sx2,sy2]],[0],True,[-1,-1],[-1,-1],self.xframesize,self.yframesizejpp, False, [0,0], levels=True, markers=markers, markercols=markercols,markersize=-2)

    def onnpp1paint(self, evt, xsize, ysize,forceredraw=True):

        if (self.nnpchecksamescale.GetValue()):
           minx=min(min(self.npp1dat[0]),min(self.npp2dat[0]))
           maxx=max(max(self.npp1dat[0]),max(self.npp2dat[0]))
           miny=min(min(self.npp1dat[1]),min(self.npp2dat[1]))
           maxy=max(max(self.npp1dat[1]),max(self.npp2dat[1]))
        else:
           minx=min(self.npp1dat[0])
           maxx=max(self.npp1dat[0])
           miny=min(self.npp1dat[1])
           maxy=max(self.npp1dat[1])

        dx=maxx-minx
        dy=maxy-miny
        minx-=0.05*dx; maxx+=0.05*dx
        miny-=0.05*dy; maxy+=0.05*dy

        dat=[]
        markers=[]

        if (self.nnpcheckfit.GetValue()):
           # a fit should be made
           istart=self.nnpfitrangesb11.GetThumbPosition()
           iend=self.nnpfitrangesb12.GetThumbPosition()
           if (iend>istart):
              [fitline,sd]=self.getnnpfitcurve(self.npp1dat[0],self.npp1dat[1],istart, iend)
              dat.append(fitline)
              markers=[[self.npp1dat[0][istart],self.npp1dat[1][istart]],[self.npp1dat[0][iend],self.npp1dat[1][iend]]]

        dat.append(self.npp1dat)

        col=[2,0]
        #print 'dat='+str(dat)

        visual.drawcurves(self.npp1panel,dat,col,False,[minx,maxx],[miny,maxy],xsize,ysize, False, [0,0], legend=[], fontsize=14, levels=True, vergrid=True, checkskip=False, linethickness=2, markers=markers,markercols=[1,1],markersize=2)
        #visual.drawcurves(self.npp1panel,dat,col,False,[-1,-1],[-1,-1],xsize,ysize, False, [0,0], legend=[], fontsize=14, levels=True, vergrid=True, checkskip=False, linethickness=2, markers=markers,markercols=[1,1],markersize=2)

    def onnpp2paint(self, evt, xsize, ysize,forceredraw=True):

        if (self.nnpchecksamescale.GetValue()):
           minx=min(min(self.npp1dat[0]),min(self.npp2dat[0]))
           maxx=max(max(self.npp1dat[0]),max(self.npp2dat[0]))
           miny=min(min(self.npp1dat[1]),min(self.npp2dat[1]))
           maxy=max(max(self.npp1dat[1]),max(self.npp2dat[1]))
        else:
           minx=min(self.npp2dat[0])
           maxx=max(self.npp2dat[0])
           miny=min(self.npp2dat[1])
           maxy=max(self.npp2dat[1])

        dx=maxx-minx
        dy=maxy-miny
        minx-=0.05*dx; maxx+=0.05*dx
        miny-=0.05*dy; maxy+=0.05*dy

        dat=[]
        markers=[]

        if (self.nnpcheckfit.GetValue()):
           # a fit should be made
           istart=self.nnpfitrangesb21.GetThumbPosition()
           iend=self.nnpfitrangesb22.GetThumbPosition()
           if (iend>istart):
              [fitline,sd]=self.getnnpfitcurve(self.npp2dat[0],self.npp2dat[1],istart, iend)
              dat.append(fitline)
              markers=[[self.npp2dat[0][istart],self.npp2dat[1][istart]],[self.npp2dat[0][iend],self.npp2dat[1][iend]]]

        dat.append(self.npp2dat)

        col=[2,0]
        #print 'dat='+str(dat)

        visual.drawcurves(self.npp2panel,dat,col,False,[minx,maxx],[miny,maxy],xsize,ysize, False, [0,0], legend=[], fontsize=14, levels=True, vergrid=True, checkskip=False, linethickness=2, markers=markers, markercols=[1,1],markersize=2)
        #visual.drawcurves(self.npp2panel,dat,col,False,[-1,-1],[-1,-1],xsize,ysize, False, [0,0], legend=[], fontsize=14, levels=True, vergrid=True, checkskip=False, linethickness=2, markers=markers, markercols=[1,1],markersize=2)

    def onnpp12paint(self, evt, xsize, ysize,forceredraw=True):

        self.onnpp1paint(wx.EVT_PAINT,xsize,ysize)
        self.onnpp2paint(wx.EVT_PAINT,xsize,ysize)

    def createoperationswindow(self,evt,ltitle,ifile):
        xsize=800
        ysize=400
        self.operationswindow=wx.Frame(self,-1, 'Dataset operations '+ltitle, (50,50), (xsize,ysize))
        self.operationswindow.SetBackgroundColour((255,255,255))

        xbuttonsize=100
        ybuttonsize=24
        dx=4
        dy=4
        dty=4

        x=dx; y=dy
        scalebutton=wx.Button(self.operationswindow, id=-1, label='Scale pattern', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        scalebutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        scalefactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        scalefactorfield.SetFont(self.textfont)
        scalefactorfield.SetValue('1.0')
        scalebutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'scale',[scalefactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(),timewindowcheckbox.GetValue()))
        scalebutton.SetToolTip(wx.ToolTip("Scale flow pattern by specified factor"))

        x=dx; y+=ybuttonsize+dy
        shiftbutton=wx.Button(self.operationswindow, id=-1, label='Shift pattern', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        shiftbutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        shiftfactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        shiftfactorfield.SetFont(self.textfont)
        shiftfactorfield.SetValue('0.0')
        shiftbutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'shift',[shiftfactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))
        shiftbutton.SetToolTip(wx.ToolTip("Shift flow pattern by specified amount"))

        x=dx; y+=ybuttonsize+dy
        whitenoisebutton=wx.Button(self.operationswindow, id=-1, label='White noise', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        whitenoisebutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        whitenoisefactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        whitenoisefactorfield.SetFont(self.textfont)
        whitenoisefactorfield.SetValue('0.0')
        whitenoisebutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'whitenoise',[whitenoisefactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))
        whitenoisebutton.SetToolTip(wx.ToolTip("Add white noise +/- the specified amount"))

        x=dx; y+=ybuttonsize+dy
        whitenoisepbutton=wx.Button(self.operationswindow, id=-1, label='White noise (%)', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        whitenoisepbutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        whitenoisepfactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        whitenoisepfactorfield.SetFont(self.textfont)
        whitenoisepfactorfield.SetValue('0.0')
        whitenoisepbutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'whitenoisep',[whitenoisepfactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))
        whitenoisepbutton.SetToolTip(wx.ToolTip("Add white noise +/- the specified percentage (relative)"))


        x=dx; y+=ybuttonsize+dy
        gaussnoisebutton=wx.Button(self.operationswindow, id=-1, label='Gaussian noise', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        gaussnoisebutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        gaussnoisefactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        gaussnoisefactorfield.SetFont(self.textfont)
        gaussnoisefactorfield.SetValue('0.0')
        gaussnoisebutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'gaussnoise',[gaussnoisefactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))
        gaussnoisebutton.SetToolTip(wx.ToolTip("Add Gaussian noise with the specified standard deviation"))

        x=dx; y+=ybuttonsize+dy
        gaussnoisepbutton=wx.Button(self.operationswindow, id=-1, label='Gaussian noise (%)', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        gaussnoisepbutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        gaussnoisepfactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        gaussnoisepfactorfield.SetFont(self.textfont)
        gaussnoisepfactorfield.SetValue('0.0')
        gaussnoisepbutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'gaussnoisep',[gaussnoisepfactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))
        gaussnoisepbutton.SetToolTip(wx.ToolTip("Add Gaussian noise with the specified standard deviation as a percentage of the local value"))

        x=dx; y+=ybuttonsize+dy
        linearbutton=wx.Button(self.operationswindow, id=-1, label='Gen. linear', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        linearbutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        startfactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        startfactorfield.SetFont(self.textfont)
        startfactorfield.SetValue('1.0')
        x+=0.5*xbuttonsize+dx
        endfactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        endfactorfield.SetFont(self.textfont)
        endfactorfield.SetValue('2.0')
        linearbutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'linear',[startfactorfield.GetValue(),endfactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))
        linearbutton.SetToolTip(wx.ToolTip("Generate linear pattern in existing data"))

        x=dx; y+=ybuttonsize+dy
        triangbutton=wx.Button(self.operationswindow, id=-1, label='Gen. triang', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        triangbutton.SetFont(self.textfont)
        x+=xbuttonsize+dx
        startfactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        startfactorfield.SetFont(self.textfont)
        startfactorfield.SetValue('1.0')
        x+=0.5*xbuttonsize+dx
        endfactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        endfactorfield.SetFont(self.textfont)
        endfactorfield.SetValue('2.0')
        triangbutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'triang',[startfactorfield.GetValue(),endfactorfield.GetValue()],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))
        triangbutton.SetToolTip(wx.ToolTip("Generate triangular pattern in existing data"))

        # pressure dependent perturbation 
        x=dx; y+=ybuttonsize+dy
        presdepbutton=wx.Button(self.operationswindow, id=-1, label='Gen. Pressd.', pos=(x, y), size=(xbuttonsize,ybuttonsize))
        presdepbutton.SetFont(self.textfont)
        presdepbutton.SetToolTip(wx.ToolTip("Generate pressure dependent perturbation on existing data"))
        x+=xbuttonsize+dx
        leakageratefield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        leakageratefield.SetFont(self.textfont)
        leakageratefield.SetValue('1')
        leakageratefield.SetToolTip(wx.ToolTip("Leakage rate for pressure factor=1"))
        x+=0.5*xbuttonsize+dx
        alphafactorfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.5*xbuttonsize,ybuttonsize))
        alphafactorfield.SetFont(self.textfont)
        alphafactorfield.SetValue('0.5')
        alphafactorfield.SetToolTip(wx.ToolTip("Exponent alpha in pressure - leakage flow rate relation"))
        x+=0.5*xbuttonsize+dx
        hourfactors=[]
        for i in range(24):
           ltc=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(0.2*xbuttonsize,ybuttonsize))
           hourfactors.append(ltc)
           l=hourfactors[-1]
           l.SetValue('1')
           label=wx.StaticText(self.operationswindow,-1,str(i),pos=(x,y-ybuttonsize+12))
           x+=0.2*xbuttonsize+0.3*dx
        presdepbutton.Bind(wx.EVT_BUTTON, lambda event: self.butpatternoperations(wx.EVT_BUTTON,'presdep',[leakageratefield.GetValue(),alphafactorfield.GetValue(),hourfactors],ifile,mintfield.GetValue(), maxtfield.GetValue(),minxfield.GetValue(), maxxfield.GetValue(), timewindowcheckbox.GetValue()))


        x=dx; y+=ybuttonsize+dy
        timewindowcheckbox=wx.CheckBox(self.operationswindow,id=-1,label='Apply time window',pos=(x,y))
        timewindowcheckbox.SetFont(self.textfont)
        timewindowcheckbox.SetValue(False)

        x=dx; y+=ybuttonsize+dy
        label=wx.StaticText(self.operationswindow,-1,'Time window:',pos=(x,y))
        label.SetFont(self.textfont)
        x+=0.8*xbuttonsize+dx; y-=dty
        mintfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(xbuttonsize-dx,ybuttonsize))
        x+=xbuttonsize+dx
        maxtfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(xbuttonsize-dx,ybuttonsize))
        mintfield.SetValue("00:00:00")
        maxtfield.SetValue("24:00:00")
        mintfield.SetFont(self.textfont)
        maxtfield.SetFont(self.textfont)

        x=dx; y+=ybuttonsize+dy+dty
        label=wx.StaticText(self.operationswindow,-1,'Period Window:',pos=(x,y))
        label.SetFont(self.textfont)
        x+=0.8*xbuttonsize+dx; y-=dty
        minxfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(2*xbuttonsize,ybuttonsize))
        x+=2*xbuttonsize+dx
        maxxfield=wx.TextCtrl(self.operationswindow,pos=(x,y),size=(2*xbuttonsize,ybuttonsize))
        minxfield.SetValue(conversion.exceldate2string(min(self.datx)))
        maxxfield.SetValue(conversion.exceldate2string(max(self.datx)))
        minxfield.SetFont(self.textfont)
        maxxfield.SetFont(self.textfont)


        self.operationswindow.Show()

    def butpatternoperations(self,evt,op,fact,ifile,tmint,tmaxt,tminx,tmaxx, timewindow):

        mint=conversion.timstr2num(tmint)
        maxt=conversion.timstr2num(tmaxt)
        minx=conversion.string2exceldate(tminx)
        maxx=conversion.string2exceldate(tmaxx)

        ifact=0
        for lfact in fact:
           try:
              if not (op=='presdep' and ifact==2):
                 dfact=float(lfact)
           except:
              dlg = wx.MessageDialog(self, 'Parameter should be a real number', 'Error', wx.ICON_ERROR)
              dlg.ShowModal()
              return
           ifact+=1

        if (ifile==1):
           ldatx=self.datx[:]
           ldaty=self.daty[:]
           ldatt=self.datt[:]
           #print 'self.daty='+str(self.daty)
        else:
           ldatx=self.datx2[:]
           ldaty=self.daty2[:]
           ldatt=self.datt2[:]
 
        odaty=[]
        n=len(ldaty)
        if op=='scale':
           for i in range(n):
              if (not timewindow) or (ldatt[i]%1>=mint) and (ldatt[i]%1<=maxt):
                 if (not timewindow) or (ldatx[i]>=minx and ldatx[i]<=maxx):
                    odaty.append(ldaty[i]*dfact)
                 else:
                    odaty.append(ldaty[i])
              else:
                 odaty.append(ldaty[i])
        elif op=='shift':
           for i in range(n):
              if (not timewindow) or (ldatt[i]%1>=mint) and (ldatt[i]%1<=maxt):
                 if (not timewindow) or (ldatx[i]>=minx) and (ldatx[i]<=maxx):
                    odaty.append(ldaty[i]+dfact)
                 else:
                    odaty.append(ldaty[i])
              else:
                 odaty.append(ldaty[i])
        elif op=='linear':
           # generate linear pattern in existing array
           ist=float(fact[0])
           ien=float(fact[1])
           for i in range(n):
              if (not timewindow) or (ldatx[i]>=minx and ldatx[i]<=maxx):
                 odaty.append(ist+(float(i)/float(n-1))*(ien-ist))
              else:
                 odaty.append(ldaty[i])
        elif op=='triang':
           # generate sawtooth pattern in existing array
           ist=float(fact[0])
           ien=float(fact[1])
           for i in range(n/2):
              if (not timewindow) or (ldatx[i]>=minx and ldatx[i]<=maxx):
                 odaty.append(ist+(float(i)/float(n/2-1))*(ien-ist))
              else:
                 odaty.append(ldaty[i])
           lasti=i
           for i in range(lasti+1,n):
              if (not timewindow) or (ldatx[i]>=minx and ldatx[i]<=maxx):
                 odaty.append(ien-(float(i-lasti)/float(n-lasti-1))*(ien-ist))
              else:
                 odaty.append(ldaty[i])
        elif op=='presdep':
           leakagerate=float(fact[0])
           alpha=float(fact[1])
           nperiod=len(fact[2])
           pfacts=[]; lfacts=[]
           for i in range(len(ldatt)):
              if (not timewindow) or (ldatx[i]>=minx and ldatx[i]<=maxx):
                 t=ldatt[i]
                 y=ldaty[i]
                 
                 iperiod=int((t%1)*float(nperiod))
                 pfact=float(fact[2][iperiod].GetValue())
                 pfacts.append(pfact)
                 lr=(pfact**alpha)*leakagerate
                 lfacts.append(lr)
                 odaty.append(y+lr)
              else:
                 odaty.append(ldaty[i])


        elif op=='whitenoise' or op=='whitenoisep':
           noiseamp=float(fact[0])
           for i in range(len(ldatt)):
              if (not timewindow) or (ldatt[i]%1>=mint) and (ldatt[i]%1<=maxt):
                 if (not timewindow) or (ldatx[i]>=minx) and (ldatx[i]<=maxx):
                    if op=='whitenoise':
                       ra=random.uniform(-noiseamp,noiseamp)
                    else:
                       l=0.01*noiseamp*ldaty[i]
                       ra=random.uniform(-l,l)
                    odaty.append(ldaty[i]+ra)
                 else:
                    odaty.append(ldaty[i])
              else:
                 odaty.append(ldaty[i])
     
        elif op=='gaussnoise' or op=='gaussnoisep':
           sd=float(fact[0])
           for i in range(len(ldatt)):
              if (not timewindow) or (ldatt[i]%1>=mint) and (ldatt[i]%1<=maxt):
                 if (not timewindow) or (ldatx[i]>=minx) and (ldatx[i]<=maxx):
                    if op=='gaussnoise':
                       ra=random.gauss(0.0, sd)
                    else:
                       ra=random.gauss(0.0, 0.01*sd*ldaty[i])
                    odaty.append(ldaty[i]+ra)
                 else:
                    odaty.append(ldaty[i])
              else:
                 odaty.append(ldaty[i])

        else:
           print('operation error')
           sys.exit(1)

        if (ifile==1):
           self.daty=odaty[:]
           self.panelredrawstatus["fppanel"]=True 
           self.panelredrawstatus["zoompanel"]=True 
           self.panelredrawstatus["histpanel"]=True 
           self.onfppaint(wx.EVT_PAINT, self.fppanel,'fppanel')
           self.onzoompaint(wx.EVT_PAINT, self.zoompanel,'zoompanel')
           self.onhistpaint(wx.EVT_PAINT, self.histpanel,'histpanel')
           self.maxyfield.SetValue(str(max(self.daty)))
           #print 'operations 1 finished'
           #print 'self.daty='+str(self.daty)
        else:
           self.daty2=odaty[:]
           self.panelredrawstatus["fppanel2"]=True 
           self.panelredrawstatus["zoompanel2"]=True 
           self.panelredrawstatus["histpanel2"]=True 
           self.onfppaint2(wx.EVT_PAINT, self.fppanel2,'fppanel2')
           self.onzoompaint2(wx.EVT_PAINT, self.zoompanel2,'zoompanel2')
           self.onhistpaint2(wx.EVT_PAINT, self.histpanel2,'histpanel2')
           self.maxyfield2.SetValue(str(max(self.daty2)))
           #print 'operations 2 finished'


    def aggregate(self,evt,tit,iset):
        pass
        dlg = wx.NumberEntryDialog(self, 'Time block duration (minutes)', '','Aggregate dataset '+str(iset),15,1,1440)
        res=dlg.ShowModal()
 
        if res==wx.ID_OK:
           nmin=dlg.GetValue()
           nday=float(nmin)/1440.

           if iset==1:
              x=self.datx[:]; y=self.daty[:]; t=self.datt[:]
              [self.datx,self.daty,self.datt]=mystat.aggregate(x,y,t,nday)
              self.panelredrawstatus["fppanel"]=True 
              self.panelredrawstatus["zoompanel"]=True 
              self.panelredrawstatus["histpanel"]=True 
              self.onfppaint(wx.EVT_PAINT,self.fppanel,"fppanel")
              self.onzoompaint(wx.EVT_PAINT, self.zoompanel, "zoompanel")
              self.onhistpaint(wx.EVT_PAINT, self.histpanel, "histpanel")
           else:
              x=self.datx2[:]; y=self.daty2[:]; t=self.datt2[:]
              [self.datx2,self.daty2,self.datt2]=mystat.aggregate(x,y,t,nday)
              self.panelredrawstatus["fppanel2"]=True 
              self.panelredrawstatus["zoompanel2"]=True 
              self.panelredrawstatus["histpanel2"]=True 
              self.onfppaint2(wx.EVT_PAINT,self.fppanel2,"fppanel2")
              self.onzoompaint2(wx.EVT_PAINT, self.zoompanel2,"zoompanel2")
              self.onhistpaint2(wx.EVT_PAINT, self.histpanel2, "histpanel2")


    def calccrosstab(self,iset,nday,periodselection,filteringtype,consistentfraction,numberofsections,numberofsectionskeep,criterion,subdivisionmethod, periodoverride=False,dataperiod=[0,0],nprior=0,returndata=False,minmaxyoverride=False, dlgparent='default'):

       minyear=-1; minmonth=-1
       if periodoverride:
          minx=supportfunctions.getPriorStartTimestamp(dataperiod[0],periodselection,nday,nprior) 
          maxx=dataperiod[1]
          if iset==1:
             if minmaxyoverride:
                miny=-1e33
                maxy=1e33
             else:
                miny=float(self.minyfield.GetValue())
                maxy=float(self.maxyfield.GetValue())
             mint=conversion.timstr2num(self.mintfield.GetValue())
             maxt=conversion.timstr2num(self.maxtfield.GetValue())
             if periodselection=="month" or periodselection=="year":
                minyear,minmonth=conversion.timstr2yearmonth(conversion.exceldate2string(dataperiod[0]))
                maxyear,maxmonth=conversion.timstr2yearmonth(conversion.exceldate2string(dataperiod[1]))
             [ldatx,ldaty,si]=mystat.subset(self.datx, self.daty, minx,maxx, miny, maxy, mint, maxt)
          else:
             if minmaxyoverride:
                miny=-1e33
                maxy=1e33
             else:
                miny=float(self.minyfield2.GetValue())
                maxy=float(self.maxyfield2.GetValue())
             mint=conversion.timstr2num(self.mintfield2.GetValue())
             maxt=conversion.timstr2num(self.maxtfield2.GetValue())
             if periodselection=="month" or periodselection=="year":
                minyear,minmonth=conversion.timstr2yearmonth(conversion.exceldate2string(dataperiod[0]))
                maxyear,maxmonth=conversion.timstr2yearmonth(conversion.exceldate2string(dataperiod[1]))
             [ldatx,ldaty,si]=mystat.subset(self.datx2, self.daty2, minx,maxx, miny, maxy, mint, maxt)

       else:
          minx=supportfunctions.getPriorStartTimestamp(self.minxzoom,periodselection,nday,nprior) 
          maxx=self.maxxzoom
          if iset==1:
             if minmaxyoverride:
                miny=-1e33
                maxy=1e33
             else:
                miny=float(self.minyfield.GetValue())
                maxy=float(self.maxyfield.GetValue())
             mint=conversion.timstr2num(self.mintfield.GetValue())
             maxt=conversion.timstr2num(self.maxtfield.GetValue())
             if periodselection=="month" or periodselection=="year":
                minyear,minmonth=conversion.timstr2yearmonth(self.minxfield.GetValue())
                maxyear,maxmonth=conversion.timstr2yearmonth(self.maxxfield.GetValue())
             [ldatx,ldaty,si]=mystat.subset(self.datx, self.daty, minx, maxx, miny, maxy, mint, maxt)
          else:
             if minmaxyoverride:
                miny=-1e33
                maxy=1e33
             else:
                miny=float(self.minyfield2.GetValue())
                maxy=float(self.maxyfield2.GetValue())
             mint=conversion.timstr2num(self.mintfield2.GetValue())
             maxt=conversion.timstr2num(self.maxtfield2.GetValue())
             if periodselection=="month" or periodselection=="year":
                minyear,minmonth=conversion.timstr2yearmonth(self.minxfield2.GetValue())
                maxyear,maxmonth=conversion.timstr2yearmonth(self.maxxfield2.GetValue())
             [ldatx,ldaty,si]=mystat.subset(self.datx2, self.daty2, minx, maxx, miny, maxy, mint, maxt)

       dat=[ldatx,ldaty]
       print('len ldatx,ldaty='+str([len(ldatx),len(ldaty)]))
       if not len(ldatx)==0:
          if iset==1:
             print('data range:'+str([min(self.datx),max(self.datx)]))
          else:
             print('data range:'+str([min(self.datx2),max(self.datx2)]))
          print('miny,maxy='+str([miny,maxy]))
          if periodoverride:
             print('minx,maxx,mint,maxt='+str([dataperiod[0],dataperiod[1],mint,maxt]))
          else:
             if iset==1:
                print('minx,maxx,mint,maxt='+str([self.minxfield.GetValue(),self.maxxfield.GetValue(),mint,maxt]))
             else:
                print('minx,maxx,mint,maxt='+str([self.minxfield2.GetValue(),self.maxxfield2.GetValue(),mint,maxt]))

       kstart=math.floor(minx)
       kend=math.ceil(maxx)


       if periodselection=="manual":
          nperiod=int(math.floor((kend-kstart)/float(nday)))
          #nperiod=int(math.floor((kend-kstart)/float(nday)))+1
       elif periodselection=="month":
          nperiod=12-minmonth+1+12*(maxyear-minyear-1)+maxmonth
       elif periodselection=="year":
          nperiod=maxyear-minyear+1
       else:
          print("calccrosstab: error periodselection")
          sys.exit(1)
       self.crosstabnperiod=nperiod

       ctslope=numpy.ones([nperiod,nperiod])
       ctintercept=numpy.zeros([nperiod,nperiod])
       ctuse=numpy.ones([nperiod,nperiod])
       ctslopesigma=numpy.zeros([nperiod,nperiod])
       ctinterceptsigma=numpy.zeros([nperiod,nperiod])

       if dlgparent=='default':
          dlgparent=self.crosstabwindow
          
       print('len(ldatx)=%i' % len(ldatx))
       if len(ldatx)>1000:
           # arbitrary boundary; wx.ProgressDialog sometimes crashes upon update for small datasets
           dlg = wx.ProgressDialog('Progress', 'Comparing periods', int(nperiod*(nperiod-1)/2),style=wx.PD_CAN_ABORT|wx.PD_ELAPSED_TIME|wx.PD_REMAINING_TIME|wx.PD_AUTO_HIDE, parent=dlgparent)
       else: 
           dlg = None;


       self.tabfitdata=[]
       for jperiod in range(nperiod):
          l=[]
          for iperiod in range(nperiod):
             l.append([])
          self.tabfitdata.append(l)

       iprogress=0
       for iperiod in range(nperiod):
          istart,iend=self.selectperiod(periodselection, iperiod, kstart,nday,minyear,minmonth)
          for jperiod in range(iperiod+1,nperiod):
             #if (iperiod==3 and jperiod==5):
                #self.verbose=True
             #else:
             self.verbose=False
             iprogress+=1
             if  dlg==None:
                abr=True
                #skip=None
             else:
                [abr,skp]=dlg.Update(iprogress)
             if not abr:
                # aborted
                dlg.Destroy()
                return False,[],[],[],[],[]
             jstart,jend=self.selectperiod(periodselection, jperiod, kstart,nday,minyear,minmonth)

             win1=[istart,iend,miny,maxy,mint,maxt]
             win2=[jstart,jend,miny,maxy,mint,maxt]
             #print 'win1='+str(win1)
             #print 'win2='+str(win2)
             try:
                lfilter=self.filteringrb_time.GetValue()
                lfrac=self.timingcombofraction
             except:
                # has not been created yet
                lfilter=False
                lfrac=100
             [result1,jpp1,jpp2,newindex]=self.calcsinglejpp(win1,win2,dat,dat)
             [result2,xxset1times,xxset2times,xxmarkercols]=self.calcsinglejppxx(win1,win2,dat,dat,timingcombofilter=lfilter,fractionperc=lfrac)
             #print 'istart,iend='+str([istart,iend])+' jstart,jend='+str([jstart,jend])
             #print 'len jpp1='+str(len(jpp1))+'   len jpp2='+str(len(jpp2))
             if len(jpp1)>1 and len(jpp2)>1 and result1 and result2:
                [fitline,sd,lusedpoints,lignoredpoints,lmarkercolors]=self.getnnpfitcurvefrac(jpp1,jpp2,xxset1times, xxset2times, 0,len(jpp1)-1,filteringtype,consistentfraction,ncluster=numberofsections,nclusterkeep=numberofsectionskeep,criterion=criterion,subdivisionmethod=subdivisionmethod,timingcombofraction=self.timingcombofraction)
                ctuse[iperiod][jperiod]=1 
                if supportfunctions.checkvalues([self.nppfitslope]) and supportfunctions.checkvalues([self.nppfitintercept]):
                   ctslope[iperiod][jperiod]=self.nppfitslope
                   ctintercept[iperiod][jperiod]=self.nppfitintercept
                   ctslopesigma[iperiod][jperiod]=self.nppfitsdslope
                   ctinterceptsigma[iperiod][jperiod]=self.nppfitsdintercept
                   self.tabfitdata[iperiod][jperiod]=[istart,iend,jstart,jend,jpp1,jpp2,fitline,sd,self.nppfitslope,self.nppfitintercept,lusedpoints,lignoredpoints,lmarkercolors,xxset1times,xxset2times,xxmarkercols]
                else:
                   ctuse[iperiod][jperiod]=0 
                   self.tabfitdata[iperiod][jperiod]=[    -1,  -1,    -1,  -1,  [],  [],     [],[],-1,-1,[],[],[],[],[],[]]
             else:
                ctuse[iperiod][jperiod]=0 
                self.tabfitdata[iperiod][jperiod]=[    -1,  -1,    -1,  -1,  [],  [],     [],[],-1,-1,[],[],[],[],[],[]]

       #print 'self.tabfitdata='+str(self.tabfitdata)

       if not dlg==None:
           dlg.Destroy()

       if returndata:
          return True,ctslope,ctintercept,ctuse,ctslopesigma, ctinterceptsigma,dat
       else:
          return True,ctslope,ctintercept,ctuse,ctslopesigma, ctinterceptsigma

    def selectperiod(self,periodselection, iperiod, kstart,nday,minyear,minmonth):
       if periodselection=="manual":
          istart=kstart+iperiod*nday
          iend=kstart+(iperiod+1)*nday
       elif periodselection=="month":
          iyear=minyear+divmod(minmonth+iperiod,12)[0]
          imonth=divmod(minmonth+iperiod,12)[1]
          if imonth==0:
             imonth=12
             iyear-=1
          #print 'year,month='+str([iyear,imonth])
          dstr="01/"+str(imonth)+"/"+str(iyear)+" 00:00:00"
          #print 'dstr='+dstr
          istart=conversion.string2exceldate(dstr)
          lastdayofmonth=calendar.monthrange(iyear,imonth)[1]
          dstr2=str(lastdayofmonth)+"/"+str(imonth)+"/"+str(iyear)+" 23:59:59"
          #print 'date range='+dstr+' - '+dstr2
          iend=conversion.string2exceldate(dstr2)
       elif periodselection=="year":
          iyear=minyear+iperiod
          dstr="01/01/"+str(iyear)+" 00:00:00"
          istart=conversion.string2exceldate(dstr)
          dstr2="31/12/"+str(iyear)+" 00:00:00"
          iend=conversion.string2exceldate(dstr2)
       else:
          print("calccrosstab: error periodselection")
          sys.exit(1)
       onesec=1.0/86400.0
       return istart,iend-onesec

    def createcrosstabwin(self,evt,iset):
       print('cross table for dataset '+str(iset))

       xsize=int(self.xsize-50)
       ysize=int(xsize/2)
       print('window size = %i x %i' % (xsize,ysize))
       #       xsize=1100
       #ysize=550
       margin1=4
       margin2=20
       xbs=80
       ybs=25
       yts=14

       self.crosstabwindow=wx.Frame(self,-1, 'Period cross tables', (50,50), (xsize,ysize))
       self.crosstabwindow.SetBackgroundColour((255,255,255))
       self.iset=iset

       xl=int((xsize-2*margin1-2*margin2)/3-8)
       #xl=(xsize-2*margin1-margin2)/2
       yl=int(xl+20)
       gyl=int(xl/5)
       #gyl=int(xl/10)

       self.crosstabxl=xl
       self.crosstabyl=yl
       self.gausspanelyl=gyl

       self.tabmarkersa=[]
       self.tabmarkersb=[]
       self.markercolsa=[]
       self.markercolsb=[]
       self.clusterreferencematrix=[]
       self.blockgroups=[]
 
       self.blockgroups1=[]
       self.clusters1=[]
       self.clusterstats1=[]
       self.clusterstatstype='distributions'
       self.blockgroups2=[]
       self.clusters2=[]
       self.clusterstats2=[]
       self.blockgroups1amp=None
       self.blockgroups2amp=None

       self.blocksteplist1=[]
       self.blocksteplist2=[]

       # filtering parameters
       self.filteringtype='none'
       self.consistentfraction=100
       self.timingcombofraction=5
       self.numberofsections=10
       self.numberofsectionskeep=8
       self.jppfitusednodes=[] # point numbers which have been used in the determination of the linear best fit
       self.clustercriterion='median a'
       self.subdivisionmethod='optimal'

       self.persel, self.crosstabnday=supportfunctions.determineBlockAnalysisDefaults(self,iset)

       nprior=0
       #nprior=3

       [success,self.slopedat,self.interceptdat,self.fitusedat,self.slopesigmadat,self.interceptsigmadat]=self.calccrosstab(iset,self.crosstabnday,self.persel,self.filteringtype,self.consistentfraction,self.numberofsections,self.numberofsectionskeep,self.clustercriterion,self.subdivisionmethod, nprior=nprior)
       if not success:
          self.crosstabwindow.Destroy()
          return
       #print 'self.slopedat size='+str(self.slopedat.shape)+' self.interceptdat size='+str(self.interceptdat.shape)

       x=margin1; y=margin1
       sllabel=wx.StaticText(self.crosstabwindow,-1,'slopes:',pos=(x,y))
       sllabel.SetFont(self.textfont)
       y+=yts
       self.slopepanellabx=x
       self.slopepanellaby=y+yl+margin1
       self.slopepanel = wx.Panel(self.crosstabwindow, pos=(x,y), size=(xl,yl))
       self.slopepanelstaticbitmap = wx.StaticBitmap(self.slopepanel,-1,pos=(0,0), size=(xl,yl))
       self.panelredrawstatus["slopepanel"]=True 
       self.slopepanelstaticbitmap.Bind(wx.EVT_LEFT_DOWN, self.OnLeftClickslopeintdn)
       self.slopepanelstaticbitmap.Bind(wx.EVT_LEFT_UP, lambda event: self.OnLeftClickslopeintup(event,1))
       self.slopepanelstaticbitmap.Bind(wx.EVT_MOTION, lambda event: self.OnLeftClickslopeintmotion(event,1))

       x+=xl+margin2;y-=yts
       iclabel=wx.StaticText(self.crosstabwindow,-1,'intercepts:',pos=(x,y))
       iclabel.SetFont(self.textfont)
       y+=yts
       self.interceptpanellabx=x
       self.interceptpanellaby=y+yl+margin1
       self.interceptpanel = wx.Panel(self.crosstabwindow, pos=(x,y), size=(xl,yl))
       self.interceptpanelstaticbitmap = wx.StaticBitmap(self.interceptpanel,-1,pos=(0,0), size=(xl,yl))
       self.panelredrawstatus["interceptpanel"]=True 

       self.interceptpanelstaticbitmap.Bind(wx.EVT_LEFT_DOWN, self.OnLeftClickslopeintdn)
       self.interceptpanelstaticbitmap.Bind(wx.EVT_LEFT_UP, lambda event: self.OnLeftClickslopeintup(event,2))
       self.interceptpanelstaticbitmap.Bind(wx.EVT_MOTION, lambda event: self.OnLeftClickslopeintmotion(event,2))

       x+=xl+margin2;y-=yts
       iclabel=wx.StaticText(self.crosstabwindow,-1,'standard deviations:',pos=(x,y))
       iclabel.SetFont(self.textfont)
       y+=yts
       self.stdevpanellabx=x
       self.stdevpanellaby=y+yl+margin1
       self.stdevpanel = wx.Panel(self.crosstabwindow, pos=(x,y), size=(xl,yl))
       self.panelredrawstatus["stdevpanel"]=True 
       self.stdevpanelstaticbitmap = wx.StaticBitmap(self.stdevpanel,-1,pos=(0,0), size=(xl,yl))

       self.stdevpanelstaticbitmap.Bind(wx.EVT_LEFT_DOWN, self.OnLeftClickslopeintdn)
       self.stdevpanelstaticbitmap.Bind(wx.EVT_LEFT_UP, lambda event: self.OnLeftClickslopeintup(event,3))

       self.slopepanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickCrosstabs)
       self.interceptpanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickCrosstabs)
       self.stdevpanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickCrosstabs)


       # data filtering control
       # options
       x=margin1; y+=yl+margin2+margin1


       # period controls
       x=margin1
       y=supportfunctions.blockAnalysisPeriodControlSet(self,self.crosstabwindow,x,y,margin1,margin2,xbs,ybs,yts,xl,yl)
       self.periodspin.Bind(wx.EVT_SPINCTRL, self.resetblockperiod)
       self.setperiodrb.Bind(wx.EVT_RADIOBUTTON, self.resetblockperiod)
       self.monperiodrb.Bind(wx.EVT_RADIOBUTTON, self.resetblockperiod)
       self.yearperiodrb.Bind(wx.EVT_RADIOBUTTON, self.resetblockperiod)
       self.priorperiodspin.Bind(wx.EVT_SPINCTRL, self.resetblockperiod)

       # visual range control
       x=margin1
       y=supportfunctions.blockAnalysisRangeControlSet(self,self.crosstabwindow,x,y,margin1,margin2,xbs,ybs,yts,xl,yl)
       self.ctwminvala.Bind(wx.EVT_TEXT,self.redrawtabsbounds)
       self.ctwmaxvala.Bind(wx.EVT_TEXT,self.redrawtabsbounds)
       self.ctwminvalb.Bind(wx.EVT_TEXT,self.redrawtabsbounds)
       self.ctwmaxvalb.Bind(wx.EVT_TEXT,self.redrawtabsbounds)
       self.ctwfixedbounds.Bind(wx.EVT_CHECKBOX,self.redrawtabsbounds)
       self.ctwblackwhite.Bind(wx.EVT_CHECKBOX,self.forceredrawtabsbounds)
#       self.showsteplines.Bind(wx.EVT_CHECKBOX,self.forceredrawtabsbounds)


       # screenshot button
       x=xsize-margin1-4*xbs
       screenshotbutton=wx.Button(self.crosstabwindow, id=-1, label='Save image', pos=(x, y), size=(xbs,ybs))
       screenshotbutton.SetFont(self.textfont)
       screenshotbutton.Bind(wx.EVT_BUTTON, lambda event: self.selectscreenshotfile(wx.EVT_BUTTON,self.crosstabwindow,xsize,yl+40))


 
       # draw contents
       self.ontablepaint(wx.EVT_PAINT,self.slopepanelstaticbitmap,self.crosstabxl,self.crosstabyl,self.slopedat,self.fitusedat,self.tabmarkersa,'slope',self.slopepanellabx,self.slopepanellaby,self.blockgroups1,"slopepanel")
       self.ontablepaint(wx.EVT_PAINT,self.interceptpanelstaticbitmap,self.crosstabxl,self.crosstabyl,self.interceptdat,self.fitusedat,self.tabmarkersb,'intercept',self.interceptpanellabx,self.interceptpanellaby,self.blockgroups2,"interceptpanel")
       self.ontablepaint(wx.EVT_PAINT,self.stdevpanelstaticbitmap,self.crosstabxl,self.crosstabyl,[self.slopesigmadat,self.interceptsigmadat],self.fitusedat,[],'standard_deviations',self.stdevpanellabx,self.stdevpanellaby,[],"stdevpanel")


       self.crosstabwindow.Show()


       # create a,b curve window after Show()
       self.createabcurvewindow(self.xsize-250,80)

       
    def forceredrawtabsbounds(self,evt):
        self.redrawtabsbounds(evt,force=True) 

    def redrawtabsbounds(self,evt,force=False):
       #if self.ctwfixedbounds.GetValue() or force:
       # used fixed bounds has been set -> update images
       self.panelredrawstatus["slopepanel"]=True 
       self.panelredrawstatus["interceptpanel"]=True 
       self.panelredrawstatus["stdevpanel"]=True 
       self.redrawtabs()

    def selectscreenshotfile(self,evt,win,xl,yl):
       #fname='crosstab_'+str(datetime.datetime.now()).replace(" ","_")+'.png'
       win.Raise() 
       fname=''
       screenshot.onTakeScreenShot(self,evt,win,xl,yl,fname)
       # bring window back to top of hierarchy
       win.Raise() 

    def resetblockperiod(self,evt):

       if self.setperiodrb.GetValue():
          # manual period length selection
          self.persel="manual"
          self.crosstabnday=int(self.periodspin.GetValue())
       elif self.monperiodrb.GetValue():
          # monthly comparison
          self.persel="month"
       elif self.yearperiodrb.GetValue():
          # yearly comparison
          self.persel="year"
       else:
          print("resetblockperiod: error: period selection")
          sys.exit(1)

       nprior=self.priorperiodspin.GetValue()

       [success,slopedat,interceptdat,fitusedat,slopesigmadat,interceptsigmadat]=self.calccrosstab(self.iset,self.crosstabnday, self.persel, self.filteringtype,self.consistentfraction,self.numberofsections,self.numberofsectionskeep,self.clustercriterion,self.subdivisionmethod, nprior=nprior)
       #[self.slopedat,self.interceptdat,self.fitusedat,self.slopesigmadat,self.interceptsigmadat]=self.calccrosstab(self.iset,self.crosstabnday, persel, self.filteringtype,self.consistentfraction,self.numberofsections,self.numberofsectionskeep,self.clustercriterion,self.subdivisionmethod)

       if not success:
          print('calccrosstab returned success=False')
          return

       self.slopedat=slopedat
       self.interceptdat=interceptdat
       self.fitusedat=fitusedat
       self.slopesigmadat=slopesigmadat
       self.interceptsigmadat=interceptsigmadat

       self.tabmarkersa=[]
       self.tabmarkersb=[]
       self.markercolsa=[]
       self.markercolsb=[]
       self.clusterreferencematrix=[]
       self.blockgroups1=[]
       self.blockgroups2=[]
       self.blockgroups1amp=None
       self.blockgroups2amp=None

       self.clusters1=[]
       self.clusterstats1=[]
       self.clusters2=[]
       self.clusterstats2=[]

#       self.clusterspin1.SetValue(-1)
#       self.clusterspin2.SetValue(-1)

#       self.panelredrawstatus["gausspanel1"]=True
#       self.panelredrawstatus["gausspanel2"]=True
       self.panelredrawstatus["slopepanel"]=True
       self.panelredrawstatus["interceptpanel"]=True
       self.panelredrawstatus["stdevpanel"]=True
       self.panelredrawstatus["abcurvepanel"]=True

       self.redrawtabs()

#       self.ongausspanelpaint(wx.EVT_PAINT,self.gausspanel1,self.crosstabxl,self.gausspanelyl,'slope',"gausspanel1")
#       self.ongausspanelpaint(wx.EVT_PAINT,self.gausspanel2,self.crosstabxl,self.gausspanelyl,'intercept',"gausspanel2")


       # if small jppwindow is present, bring it to front and update
       try:
          # check if window if present and if so, get coordinates
          x,y=self.smalljppwindow.GetPosition()
          # replace with new window (new data) at same location
          self.smalljppwindow.Destroy()
          self.createsmalljppwindow(x,y,self.ix,self.iy)
          #self.jppwindow.Raise() # rbing to top
       except:
          pass

       self.crosstabwindow.Raise() 

    def OnLeftClickslopeintdn(self,evt):
       x, y = evt.GetPosition()
       self.startix=int(x/(self.crosstabxl/float(self.crosstabnperiod+1)))-1 # extra column for labels
       self.startiy=int(y/((self.crosstabyl-20)/float(self.crosstabnperiod+1)))-1 # extra row for labels

    def OnLeftClickslopeintmotion(self,evt,iframe):
       x, y = evt.GetPosition()
       ix=int(x/(self.crosstabxl/float(self.crosstabnperiod+1)))-1 # extra column for labels
       iy=int(y/((self.crosstabyl-20)/float(self.crosstabnperiod+1)))-1 # extra row for labels
       if evt.Dragging():
          if not (ix==self.startix and iy==self.startiy):
             # an area has been selected
             #print 'area selected: '+str([self.startix,self.startiy,ix,iy])
             if iframe==1:
                self.blockgroups1=[self.startix,self.startiy,ix,iy]
                self.blockgroups1amp=None
                self.panelredrawstatus["slopepanel"]=True # redraw to mark selected block
                self.redrawtabs()
             elif iframe==2:
                self.blockgroups2=[self.startix,self.startiy,ix,iy]
                self.blockgroups2amp=None
                self.panelredrawstatus["interceptpanel"]=True # redraw to mark selected block
                self.redrawtabs()
             else:
                # do nothing
                pass

    def OnLeftClickslopeintup(self,evt,iframe):
       x, y = evt.GetPosition()
       self.ix=int(x/(self.crosstabxl/float(self.crosstabnperiod+1)))-1 # extra column for labels
       self.iy=int(y/((self.crosstabyl-20)/float(self.crosstabnperiod+1)))-1 # extra row for labels
       #self.ix=int(x/(self.crosstabxl/float(self.crosstabnperiod))) 
       #self.iy=int(y/((self.crosstabyl-20)/float(self.crosstabnperiod))) 
       #print 'left click @ '+str([x,y])+' index '+str([self.ix,self.iy])

       if self.ix==self.startix and self.iy==self.startiy:
          # start and end of click on same cell -> show info window for this cell
          if self.iy>self.ix:
             #print 'remove marker'
             # outside area -> remove marker and destry smalljppwindow if present
             try:
                posx,posy=self.smalljppwindow.Destroy()
             except:
                pass
             self.ix=-1
             self.iy=-1
             self.tabmarkersa=[]
             self.tabmarkersb=[]
             posx=self.xsize-512 #875
             posy=300# 750
          else:
             try:
                # if a smalljppwindow already exists, copy its location for the new one
                posx,posy=self.smalljppwindow.GetPosition()
             except:
                # if not, set default values
                posx=self.xsize-510# 875
                posy=300# 750
             if iframe==1:
                # reset
                self.blockgroups1=[]
                self.blockgroups1amp=None
             elif iframe==2:
                # reset
                self.blockgroups2=[]
                self.blockgroups2amp=None
          self.panelredrawstatus["slopepanel"]=True # redraw to mark selected block
          self.panelredrawstatus["interceptpanel"]=True # redraw to mark selected block
          self.panelredrawstatus["smalljpppanel"]=True
          self.panelredrawstatus["smalljppxxpanel"]=True
          self.panelredrawstatus["abcurvepanel"]=True
          if not self.iy>self.ix:
             self.createsmalljppwindow(posx,posy,self.ix,self.iy)
   
             # window might not exist - catch error if it does not
             try:
                self.onabcurvepaint(wx.EVT_PAINT,self.abcurvepanelxl,self.abcurvepanelyl)
             except:
                pass


    def createabcurvewindow(self,posx,posy):
       res=self.tabfitdata

       if len(res)>0:
          xsize=200
          ysize=200
          self.abcurvewindow=wx.Frame(self.crosstabwindow,-1, 'a vs b', (posx,posy), (xsize,ysize),style=wx.FRAME_FLOAT_ON_PARENT|wx.CLOSE_BOX|wx.CAPTION)
          self.abcurvewindow.SetBackgroundColour((255,255,255))
    
          margin1=4

          xl=xsize-2*margin1
          yl=xl
          self.abcurvepanelxl=xl
          self.abcurvepanelyl=yl
    
          x=margin1; y=margin1
          self.abcurvepanel = wx.Panel(self.abcurvewindow, pos=(x, y), size=(xl,yl)) 
          self.abcurvepanelstaticbitmap = wx.StaticBitmap(self.abcurvepanel)
          self.panelredrawstatus["abcurvepanel"]=True
          self.abcurvepanel.Bind(wx.EVT_PAINT, lambda event: self.onabcurvepaint(wx.EVT_PAINT,xl,yl))
    
          self.abcurvepanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickabcurvepanel)
          
          self.abcurvewindow.Show()
       else:
          try:
             self.abcurvewindow.Destroy()
          except:
             pass
 

       self.crosstabwindow.Raise() 
   
    def onabcurvepaint(self,evt,xl,yl,panelname='abcurvepanel',display=True,usetabmarkercolors=True):
       if self.panelredrawstatus[panelname]:
          alist=[]; blist=[] # for plotting
          xlist=[]; ylist=[] # for R2 determination
    
          nx=len(self.tabfitdata[0])
          ny=len(self.tabfitdata)
    
          mina=1e33; maxa=-1e33; minb=1e33; maxb=-1e33
          i=-1; markeditem=-1
          for iy in range(ny):
             for ix in range(iy+1,nx):
                i+=1
                a=self.tabfitdata[iy][ix][8]
                b=self.tabfitdata[iy][ix][9]
                use=self.fitusedat[iy][ix]
                alist.append([a])
                blist.append([b])
                xlist.append(a)
                ylist.append(b)
                if use:
                   mina=min(a,mina)
                   maxa=max(a,maxa)
                   minb=min(b,minb)
                   maxb=max(b,maxb)
    
                if iy==self.iy and ix==self.ix:
                   markeditem=i

          self.blockDiagramValueBounds=[mina,maxa,minb,maxb]
    
          # determine R^2 
    
          # http://en.wikipedia.org/wiki/Simple_linear_regression
          # with asymptotic assumption and q=1 because we just want to have the standard deviation
          ym=numpy.mean(ylist)
          syym=0.0
          for i in range(len(alist)):
             syym+=(ylist[i]-ym)**2
             #print 'a,b='+str(xlist[i])+' '+str(ylist[i])
          # following http://en.wikipedia.org/wiki/Coefficient_of_determination
          #print 'xlist='+str(xlist)
          #print 'ylist='+str(ylist)
          [fit,residuals, rank, singular_values, rcond]=numpy.polyfit(xlist,ylist,1,full=True) # linear fit
    
          #print 'ym,syym,residuals[0],R2='+str([ym,syym,residuals[0],R2])
    
          try:
             R2=1.0-residuals[0]/syym
          except:
             R2=9
          if display:
             self.abcurvewindow.SetTitle('a vs b; R^2='+'%2.2f' % R2)
    
          self.abcurveR2=R2
          self.abcurvedata=[alist,blist]
    
          #print 'markeditem='+str(markeditem)+' ix,iy,nx,ny='+str([ix,iy,nx,ny])
          if usetabmarkercolors:
             markercolsa=self.markercolsa[:]
             markercolsb=self.markercolsb[:]
          else:
             markercolsa=[]
             markercolsb=[]
             for i in range(len(alist)):
                markercolsa.append(0)
                markercolsb.append(0)
             # add marked point to end of lists so that it will be shown on top
             # print 'alist='+str(alist)
             if markeditem>-1:
                alist.append(alist[markeditem])
                blist.append(blist[markeditem])
                markercolsa.append((0,255,255))
                markercolsb.append((0,255,255))
    
    
          da=0.1*(maxa-mina)
          db=0.1*(maxb-minb)
          if display:
             self.panelbuffers[panelname],outranges=visual.drawsimplepointcloud(self.abcurvepanelstaticbitmap,alist,blist,[mina-da,maxa+da],[minb-db,maxb+db],xl,yl,markercols=markercolsa,display=display,drawonetooneline=False)
          else:
             self.panelbuffers[panelname],outranges=visual.drawsimplepointcloud([],alist,blist,[mina-da,maxa+da],[minb-db,maxb+db],xl,yl,markercols=markercolsb,display=display,drawonetooneline=False,markersize=10)
          self.panelredrawstatus[panelname]=False
    
    
       else:
          visual.drawBitmap(self.abcurvepanelstaticbitmap,self.panelbuffers[panelname])


    def createsmalljppwindow(self,posx,posy,ix,iy):
       print('createsmalljppwindow')

       res=self.tabfitdata[iy][ix]

       if len(res)==16:
          [istart,iend,jstart,jend,jpp1,jpp2,fitline,sd,afact,bfact,usedpoints,ignoredpoints,markercolors,xxtimesset1,xxtimesset2,xxmarkercols]=res
          #print 'istart,iend,jstart,jend='+str([istart,iend,jstart,jend])

          # marker -> redraw matrices
          self.tabmarkersa=[[ix,iy]]
          self.tabmarkersb=[[ix,iy]]
          self.redrawtabs()

          try:
             self.smalljppwindow.Destroy()
          except:
             pass

          xsize=512
          ysize=270+64
          self.smalljppwindow=wx.Frame(self.crosstabwindow,-1, 'CFPD diagram and timing combinations', (posx,posy), (xsize,ysize),style=wx.FRAME_FLOAT_ON_PARENT|wx.CLOSE_BOX|wx.CAPTION)
          self.smalljppwindow.SetBackgroundColour((255,255,255))
  
          margin1=4
          margin2=20


          xl=int((xsize-3*margin1)/2-16)
          yl=xl

          t1=conversion.exceldate2stringd(istart)+' - '+conversion.exceldate2stringd(iend)
          t2=conversion.exceldate2stringd(jstart)+' - '+conversion.exceldate2stringd(jend)
 
          x=margin1; y=margin1
          label=wx.StaticText(self.smalljppwindow,-1,'hori: '+t1+'   vert: '+t2,pos=(x,y+4))
          font = wx.Font(5, wx.SWISS, wx.NORMAL, wx.NORMAL, False)
          label.SetFont(font)
          y+=margin2
          self.smalljpppanel = wx.Panel(self.smalljppwindow, pos=(x, y), size=(xl,yl)) 
          self.smalljpppanelstaticbitmap = wx.StaticBitmap(self.smalljpppanel)
          self.panelredrawstatus["smalljpppanel"]=True
          self.smalljpppanel.Bind(wx.EVT_PAINT, lambda event: self.onsmalljpppaint(wx.EVT_PAINT,xl,yl,jpp1,jpp2,fitline,sd,afact,bfact,markerpoints=usedpoints,ignoredpoints=ignoredpoints,markercolors=markercolors))
          
          x+=margin1+xl; y=margin1
          label=wx.StaticText(self.smalljppwindow,-1,'time combinations',pos=(x,y+4))
          y+=margin2
          self.smalljppxxpanel = wx.Panel(self.smalljppwindow, pos=(x, y), size=(xl,yl)) 
          self.smalljppxxpanelstaticbitmap = wx.StaticBitmap(self.smalljppxxpanel)
          self.panelredrawstatus["smalljppxxpanel"]=True
          self.smalljppxxpanel.Bind(wx.EVT_PAINT, lambda event: self.onsmalljppxxpaint(wx.EVT_PAINT,xl,yl,xxtimesset1,xxtimesset2,markercols=xxmarkercols))


          self.smalljppwindow.Show()
       else:
          try:
             self.smalljppwindow.Destroy()
             self.tabmarkersa=[]
             self.tabmarkersb=[]
             self.redrawtabs()
          except:
             pass
 

    def redrawtabs(self,evt=wx.EVT_PAINT):
       if self.blockgroups1amp==None:
          label1='' 
       else:
          label1='amplitude=%5.3f' % self.blockgroups1amp   
       if self.blockgroups2amp==None:
          label2='' 
       else:
          label2='amplitude=%5.3f' % self.blockgroups2amp   

 
       self.ontablepaint(wx.EVT_PAINT,self.slopepanelstaticbitmap,self.crosstabxl,self.crosstabyl,self.slopedat,self.fitusedat,self.tabmarkersa,'slope',self.slopepanellabx,self.slopepanellaby,self.blockgroups1,"slopepanel", markercols=self.markercolsa,comment=label1)
       self.ontablepaint(wx.EVT_PAINT,self.interceptpanelstaticbitmap,self.crosstabxl,self.crosstabyl,self.interceptdat,self.fitusedat,self.tabmarkersb,'intercept',self.interceptpanellabx,self.interceptpanellaby,self.blockgroups2,"interceptpanel", markercols=self.markercolsb,comment=label2)
       self.ontablepaint(wx.EVT_PAINT,self.stdevpanelstaticbitmap,self.crosstabxl,self.crosstabyl,[self.slopesigmadat,self.interceptsigmadat],self.fitusedat,[],'standard_deviations',self.stdevpanellabx,self.stdevpanellaby,[],"stdevpanel", markercols=[])

       # also update abcurve window if present:
       try:
          self.onabcurvepaint(wx.EVT_PAINT,self.abcurvepanelxl,self.abcurvepanelyl)
       except:
          pass

    def findcoherentblocks(self,dat,use,factor,dlg,idlg,minbs):
        n=dat.shape[0]

        results=[]
        iys=0; # all relevant blocks start at y=0!
        for ixs in range(2,n): # start node in x direction
           idlg+=1
           if idlg%10==0:
              [abr,skp]=dlg.Update(idlg,'Searching coherent clusters')
              if not abr:
                 # abort 
                 return True,[],[],idlg
  
           bsy=ixs # block from top to diagonal
           for bsx in range(minbs,n-ixs+1): # block size in x direction
              bi=self.subarray(dat,ixs,ixs+bsx-1,iys,iys+bsy-1) # block values
              ui=self.subarray(use,ixs,ixs+bsx-1,iys,iys+bsy-1) # block node use
              
              vals=numpy.multiply(bi,ui) # elementwise multiplication

              if int(numpy.sum(ui))==numpy.size(ui):
                 mean=numpy.mean(vals)
                 sd=numpy.std(vals)
                 results.append([ixs,iys,bsx,bsy,len(vals),mean,sd])



        results.sort(key=self.listitemabs5)
        results.reverse()


        [abort,clusters,odlg]=self.findsameblocks(results, 0.025,factor,dlg,idlg,dat)
          

        #print 'findcoherentblocks exit idlg='+str(idlg)
        return abort,results,clusters, odlg

    def listitemabs5(self,l):
        return abs(l[5]) 


    def subarray(self,arr,ixs,ixe,iys,iye):
        oa=[]
        for iy in range(iys,iye+1):
           oa.append(arr[iy][ixs:ixe+1])
        return oa

   
    def keyfunc1(self,*items):
        #print 'items='+str(items)
        if len(items[0])>=3:
           if self.analyzerankingmode=='mean/stdev':
              s=-abs(items[0][1]/items[0][2]) # largest deviation of mean divided by standard deviation first
           elif self.analyzerankingmode=='mean':
              s=-abs(items[0][1]) # largest deviation of mean first
           elif self.analyzerankingmode=='stdev':
              s=items[0][2] # stdev
           elif self.analyzerankingmode=='stdev/n':
              s=items[0][2]/float(len(items[0][0])) # stdev divided by number of blocks in group
           elif self.analyzerankingmode=='mean/n_stdev':
              s=abs(items[0][1])/(items[0][2]*float(len(items[0][0]))) # mean divided by number of blocks in group times stdev
           elif self.analyzerankingmode=='n_mean/stdev':
              s=-abs(items[0][1])/(items[0][2])*float(len(items[0][0])) # mean divided by number of blocks in group times stdev
           elif self.analyzerankingmode=='block size':
              s=-float(len(items[0][0]))
           else:
              print('undefined self.analyzerankingmode "'+self.analyzerankingmode+'"')
              sys.exit(1)
           return s
        else:
           return 1e33

    def uniqify(self,dat):
        ldict={}
        for item in dat:
           ldict[str(item)]=1
        llist=[]
        #print 'ldict='+str(ldict)
        if len(ldict)>0:
           lkeys=ldict.keys()
           lkeys.sort() 
        #   print 'lkeys='+str(lkeys)
           for item in lkeys:
              s=item.replace('[','').replace(']','').split(',')
              ll=[]
              for jtem in s:
                 ll.append(int(jtem))
              llist.append(ll)

        return llist


    def onsmalljpppaint(self,evt,xl,yl,jpp1,jpp2,fitline,sd,a,b,markerpoints=[],ignoredpoints=[],markercolors=[],panelname="smalljpppanel"):

        if self.panelredrawstatus[panelname]:
           minxy=min(min(jpp1),min(jpp2))
           maxxy=max(max(jpp1),max(jpp2))

           dxy=maxxy-minxy
           minxy-=0.05*dxy; maxxy+=0.05*dxy

           dat=[[[minxy,maxxy],[minxy,maxxy]]] # equal probability line
           lsd=[[]]

           dat.append(fitline)
           lsd.append(sd)
           col=[1,2,0]

           dat.append([jpp1,jpp2])
           lsd.append([])

           fontsize=8

           lstr='a=%5.3f' %a
           lstr+='\nb=%5.3f' %b
           #print 'lstr='+lstr

           markers=[]
           markercols=[]
           i=0
           for item in markerpoints:
              markers.append([jpp1[item],jpp2[item]])
              try: 
                 markercols.append(markercolors[i])
              except:
                 markercols.append(2)
              i+=1
           for item in ignoredpoints:
              markers.append([jpp1[item],jpp2[item]])
              markercols.append(1)
           markersize=-2
           #print 'markercolors='+str(markercolors)
           #print 'markers='+str(markers)
           #print 'markercols='+str(markercols)

           self.panelbuffers[panelname]=visual.drawcurves(self.smalljpppanelstaticbitmap,dat,col,False,[minxy,maxxy],[minxy,maxxy],xl,yl, False, [0,0], sd=lsd,legend=[], fontsize=fontsize, levels=True, vergrid=True, checkskip=False, linethickness=2,labelgridline=True, horvergridsame=True,text=lstr,markers=markers,markersize=markersize,markercols=markercols)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(self.smalljpppanel,self.panelbuffers[panelname])

    def onsmalljppxxpaint(self,evt,xl,yl,x,y,panelname="smalljppxxpanel",markercols=[]):

        if self.panelredrawstatus[panelname]:

           self.panelbuffers[panelname],outranges=visual.drawsimplepointcloud(self.smalljppxxpanelstaticbitmap,x,y,[0.0,1.0],[0.0,1.0],xl,yl,markercols=markercols,drawonetooneline=False)
           #print 'x='+str(x)
           #print 'y='+str(y)
           self.panelredrawstatus[panelname]=False
        else:
           visual.drawBitmap(self.smalljppxxpanelstaticbitmap,self.panelbuffers[panelname])



    def ontablepaint(self,evt,staticbitmap,xl,yl,dat,use,markers,label,labx,laby,clusters,panelname,display=True,printvaluerange=False,markercols=[],comment=''):
       print('ontablepaint')
       if self.panelredrawstatus[panelname]:
          print('redraw')
          if self.ctwfixedbounds.GetValue():
             # manually set bounds
             try:
                if label=='slope':
                   amin=float(self.ctwminvala.GetValue())
                   amax=float(self.ctwmaxvala.GetValue())
                   bounds=[amin,amax]
                elif label=='intercept':
                   bmin=float(self.ctwminvalb.GetValue())
                   bmax=float(self.ctwmaxvalb.GetValue())
                   bounds=[bmin,bmax]
                else:
                   bounds=[-1,-1]
             except:
                bounds=[-1,-1]
          else:
             bounds=[-1,-1]
   
          #print 'ontablepaint matrix '+label+' size='+str(dat.shape)
          vertlines=[]
          if label=='slope':
             zerov=1.0
#             if self.showsteplines.GetValue():
#                vertlines=self.blocksteplist1
          else:
             zerov=0.0
#             if label=='intercept' and self.showsteplines.GetValue():
#                vertlines=self.blocksteplist2
    
          #print 'zerov='+str(zerov)
          if self.monperiodrb.GetValue():
             periodleg='monthly'
          else:
             periodleg='datedays'
             #periodleg='numbers'

          nprior= self.priorperiodspin.GetValue()
          bw=self.ctwblackwhite.GetValue() 
          istart=math.floor(self.minxzoom)
          self.crosstabnday=self.periodspin.GetValue()
          if self.setperiodrb.GetValue() and self.crosstabnday==1:
             singledays=True
          else:
             singledays=False
          print('nprior,bw,istart,crosstabnday,singledays='+str([nprior,bw,istart,self.crosstabnday,singledays]))
   
          [minv,maxv,self.panelbuffers[panelname]]=visual.drawMatrix(staticbitmap,dat,use,0,0,xl,yl,region='U',zerov=zerov,markers=markers,bw=bw, bounds=bounds,legend=periodleg,display=display,printvaluerange=printvaluerange,nprior=nprior,start=istart,singledays=singledays,markercols=markercols,markerbox=clusters,label=comment,vertlines=vertlines)
          self.panelredrawstatus[panelname]=False
   
          if self.ctwfixedbounds.GetValue() and not label=='standard_deviations':
             # bounds are set
             minv=bounds[0]
             maxv=bounds[1]
   
   
          if display:
             try:
                # single value
                lstr1=str(("%4.2f" % minv))
                lstr2=str(("%4.2f" % maxv))
             except:
                # two values
                lstr1=str(("%4.2f" % minv[0]))
                lstr2=str(("%4.2f" % maxv[0]))
                lstr3=str(("%4.2f" % minv[1]))
                lstr4=str(("%4.2f" % maxv[1]))
      
             if self.appsize=='large':
                textfont = wx.Font(12, wx.SWISS, wx.NORMAL, wx.NORMAL, False)
                textfontsmall = wx.Font(9, wx.SWISS, wx.NORMAL, wx.NORMAL, False)
                tdx=40
             else:
                textfont = wx.Font(8, wx.SWISS, wx.NORMAL, wx.NORMAL, False)
                textfontsmall = wx.Font(6, wx.SWISS, wx.NORMAL, wx.NORMAL, False)
                tdx=24
             if label=='slope':
                try:
                   self.matrange1a.Destroy()
                   self.matrange1b.Destroy()
                   self.matrange1c.Destroy()
                except:
                   pass
                self.matrange1a=wx.StaticText(self.crosstabwindow,-1,lstr1,pos=(labx,laby))
                x=xl*((1.0-minv)/(maxv-minv))+labx
                print('x,minv,maxv,xl,labx='+str([x,minv,maxv,xl,labx]))
                self.matrange1b=wx.StaticText(self.crosstabwindow,-1,'1',pos=(x,laby))
                self.matrange1c=wx.StaticText(self.crosstabwindow,-1,lstr2,pos=(labx+xl-tdx,laby))
                self.matrange1a.SetFont(textfont)
                self.matrange1b.SetFont(textfontsmall)
                self.matrange1c.SetFont(textfont)
             elif label=='intercept':
                try:
                   self.matrange2a.Destroy()
                   self.matrange2b.Destroy()
                   self.matrange2c.Destroy()
                except:
                   pass
                self.matrange2a=wx.StaticText(self.crosstabwindow,-1,lstr1,pos=(labx,laby))
                x=xl*((0.0-minv)/(maxv-minv))+labx
                self.matrange2b=wx.StaticText(self.crosstabwindow,-1,'0',pos=(x,laby))
                self.matrange2c=wx.StaticText(self.crosstabwindow,-1,lstr2,pos=(labx+xl-tdx,laby))
                self.matrange2a.SetFont(textfont)
                self.matrange2b.SetFont(textfontsmall)
                self.matrange2c.SetFont(textfont)
             elif label=='standard_deviations':
                try:
                   self.matrange3a.Destroy()
                   #self.matrange3b.Destroy()
                   self.matrange3c.Destroy()
                   self.matrange3d.Destroy()
                   #self.matrange3e.Destroy()
                   self.matrange3f.Destroy()
                except:
                   pass
                self.matrange3a=wx.StaticText(self.crosstabwindow,-1,lstr1,pos=(labx,laby))
                x=0.5*(xl-10)*((0.0-minv[0])/(maxv[0]-minv[0]))+labx
                #self.matrange3b=wx.StaticText(self.crosstabwindow,-1,'0',pos=(x,laby))
                self.matrange3c=wx.StaticText(self.crosstabwindow,-1,lstr2,pos=(labx+(xl-10)/2-tdx,laby))
                self.matrange3a.SetFont(textfont)
                #self.matrange3b.SetFont(textfont)
                self.matrange3c.SetFont(textfont)
                self.matrange3d=wx.StaticText(self.crosstabwindow,-1,lstr3,pos=(labx+10+(xl-10)/2,laby))
                x=0.5*(xl-10)*((0.0-minv[1])/(maxv[1]-minv[1]))+labx+xl/2
                #self.matrange3e=wx.StaticText(self.crosstabwindow,-1,'0',pos=(x,laby))
                self.matrange3f=wx.StaticText(self.crosstabwindow,-1,lstr4,pos=(labx+xl-tdx,laby))
                self.matrange3d.SetFont(textfont)
                #self.matrange3e.SetFont(textfont)
                self.matrange3f.SetFont(textfont)
       elif display:
          print('self.panelbuffers["'+panelname+'"]='+str(self.panelbuffers[panelname]))
          visual.drawBitmap(staticbitmap,self.panelbuffers[panelname])
          
       if display:
          print('setting bitmap %s' % panelname)
          print('xl=%i yl=%i'% (int(xl),int(yl)))
          print('panel %s' % str(self.panelbuffers[panelname]))
          if panelname=='slopepanel':  
             self.slopepanelstaticbitmap.SetBitmap(self.panelbuffers[panelname])
          elif panelname=='interceptpanel':    
             self.interceptpanelstaticbitmap.SetBitmap(self.panelbuffers[panelname])
          elif panelname=='stdevpanel':    
             self.stdevpanelstaticbitmap.SetBitmap(self.panelbuffers[panelname])
          

             
   
    def presssavedata(self,evt,iset):

       mint=0.0; maxt=1.0

       if iset==1:
          miny=float(self.minyfield.GetValue())
          maxy=float(self.maxyfield.GetValue())
          [sx,sy,si]=mystat.subset(self.datx, self.daty, self.minxzoom, self.maxxzoom, miny, maxy, mint, maxt)
       else:
          miny=float(self.minyfield2.GetValue())
          maxy=float(self.maxyfield2.GetValue())
          [sx,sy,si]=mystat.subset(self.datx2, self.daty2, self.minxzoom2, self.maxxzoom2, miny, maxy, mint, maxt)

       print('len(sx)='+str(len(sx))+'len(sy)='+str(len(sy)))
 
       #path = wx.SaveFileSelector('Modified data file', '.csv', '')
       dlg = wx.FileDialog(self,message='Save modified data file', defaultDir=self.cwd, wildcard='*.csv', style=wx.FD_SAVE)
       if dlg.ShowModal() == wx.ID_OK:
          path=dlg.GetPath()
          self.cwd=fsconfig.strippath(path)
          fsconfig.saveconfig([self.cwd,self.appsize])
       dlg.Destroy()
       if not (path==''):
          f=open(path,'w')
          for i in range(len(sx)):
             f.write(str(sx[i])+','+str(sy[i])+'\n')
          f.close()

    def resetconsistentfraction(self,evt):
       self.consistentfraction=int(self.consistentfractionspin.GetValue())
       # update tables
       self.resetblockperiod(wx.EVT_SPINCTRL)

    def resetnumberofsections(self,evt):
       self.numberofsections=int(self.numberofsectionsspin.GetValue())
       # update tables
       self.resetblockperiod(wx.EVT_SPINCTRL)

    def resetnumberofsectionskeep(self,evt):
       self.numberofsectionskeep=int(self.numberofsectionskeepspin.GetValue())
       # update tables
       self.resetblockperiod(wx.EVT_SPINCTRL)
         
    def resetclustercriterion(self,evt):
       self.clustercriterion=self.clustercriterioncombo.GetValue()
       # update tables
       self.resetblockperiod(wx.EVT_COMBOBOX)
         
    def resettimingcombofraction(self,evt):
       self.timingcombofraction=self.timingcombofractionspin.GetValue()
       # update tables
       self.resetblockperiod(wx.EVT_SPINCTRL)
         
    def resetsubdivisionmethod(self,evt):
       self.subdivisionmethod=self.subdivisionmethodcombo.GetValue()
       # update tables
       self.resetblockperiod(wx.EVT_COMBOBOX)
         
         
    def filteringchanged(self,evt,newt):
       self.filteringtype=newt
       # update tables
       self.resetblockperiod(wx.EVT_RADIOBUTTON)

    def resetanalyzerankingmode(self,evt):
       self.analyzerankingmode=self.analyzerankingmodecombo.GetValue()
       # redo analysis
       self.autofindblocks(wx.EVT_BUTTON)
       # redraw tables
       self.panelredrawstatus["slopepanel"]=True
       self.panelredrawstatus["interceptpanel"]=True
       self.redrawtabs()

    def OnRightClickabcurvepanel(self,evt):
       x, y = evt.GetPosition()
       
       abcurvepanelmenu=wx.Menu()
       ID_SAVEIM=101
       ID_SAVEDA=102
       abcurvepanelmenu.Append(ID_SAVEIM, "Save image", "Save image")
       abcurvepanelmenu.Append(ID_SAVEDA, "Save data", "Save data")
       self.Bind(wx.EVT_MENU, lambda event: self.abcurvepanelSaveImage(wx.EVT_MENU,self.panelbuffers['abcurvepanel']), id=ID_SAVEIM)       
       self.Bind(wx.EVT_MENU, lambda event: self.abcurvepanelSaveData(wx.EVT_MENU,self.abcurvedata) , id=ID_SAVEDA)       

       self.abcurvewindow.PopupMenu( abcurvepanelmenu, evt.GetPosition() )

       abcurvepanelmenu.Destroy() # to prevent memory leak

    def abcurvepanelSaveImage(self,evt,bitmap,parent="crosstabwindow"):
        dlg = wx.FileDialog(self,message='Save image', defaultDir=self.cwd, wildcard='*.png', style=wx.FD_SAVE)
        if dlg.ShowModal() == wx.ID_OK:
           path=dlg.GetPath()
          
           if not (path==''):
              screenshot.savebitmap(bitmap,path)
              #screenshot.savebitmap(self.panelbuffers['abcurvepanel'],path)
              self.cwd=fsconfig.strippath(path)
              fsconfig.saveconfig([self.cwd,self.appsize])
           else:
              dlg = wx.MessageDialog(self, 'Image write error', 'Error', wx.ICON_ERROR)
              dlg.ShowModal()
        else:
           dlg = wx.MessageDialog(self, 'Image write error', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()

        # bring window back to top of hierarchy
        if parent=="crosstabwindow":
           self.crosstabwindow.Raise()
        else:
           parent.Raise()

    def abcurvepanelSaveData(self,evt,data):
        dlg = wx.FileDialog(self,message='Save data', defaultDir=self.cwd, wildcard='*.csv', style=wx.FD_SAVE)
        if dlg.ShowModal() == wx.ID_OK:
           path=dlg.GetPath()
          
           if not (path==''):
              # save data
              self.safedata(path,data)
              self.cwd=fsconfig.strippath(path)
              fsconfig.saveconfig([self.cwd,self.appsize])
           else:
              dlg = wx.MessageDialog(self, 'Data write error', 'Error', wx.ICON_ERROR)
              dlg.ShowModal()
        else:
           dlg = wx.MessageDialog(self, 'Data write error', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()

        # bring window back to top of hierarchy
        self.crosstabwindow.Raise()


    def savedata(self,path,data):
        # save data
        f=open(path,'w')
        for i in range(len(data[0])):
           f.write(str(data[0][i][0])+','+str(data[1][i][0])+'\n') # every entry is a 1- or 2-item list
        f.close()

    def savedata2(self,path,data):
        # save data
        f=open(path,'w')
        for i in range(len(data[0])):
           f.write(str(data[0][i])+','+str(data[1][i])+'\n') # every entry is a 1- or 2-item list
        f.close()

    def OnRightClickCrosstabs(self,evt):
        x, y = evt.GetPosition()
       
        abspanelmenu=wx.Menu()
        ID_SAVEDA=111
        abspanelmenu.Append(ID_SAVEDA, "Save data (xls)", "Save data (xls)")
        self.Bind(wx.EVT_MENU, self.saveABSData)
 
        self.crosstabwindow.PopupMenu( abspanelmenu, evt.GetPosition() )
 
        abspanelmenu.Destroy() # to prevent memory leak

    def saveABSData(self,evt):
        dlg = wx.FileDialog(self,message='Save all table data', defaultDir=self.cwd, wildcard='*.xls', style=wx.FD_SAVE)

        if self.iset==1:
           datafile=self.datafilename1
           mint=self.minxzoom
        else:
           datafile=self.datafilename2
           mint=self.minxzoom2

        if dlg.ShowModal() == wx.ID_OK:
           path=dlg.GetPath()
  
           path.replace('.xls','').replace('.XLS','')+'.xls'
          
           if not (path==''):
              flowstatexports.exportABSToXls(self.slopedat,self.interceptdat,self.slopesigmadat,self.interceptsigmadat,mint,self.crosstabnday,self.crosstabnperiod,datafile,path)
           else:
              dlg = wx.MessageDialog(self, 'XLS write error', 'Error', wx.ICON_ERROR)
              dlg.ShowModal()
        else:
           dlg = wx.MessageDialog(self, 'XLS write error', 'Error', wx.ICON_ERROR)
           dlg.ShowModal()

        # bring window back to top of hierarchy
        self.crosstabwindow.Raise()

    def showBlockFunctions(self,evt):
       posx=20; posy=20
       xsize=940
       ysize=680
       self.blockfunctionswindow=wx.Frame(self.crosstabwindow,-1, 'Block functions', (posx,posy), (xsize,ysize),style=wx.FRAME_FLOAT_ON_PARENT|wx.CLOSE_BOX|wx.CAPTION)
       self.blockfunctionswindow.SetBackgroundColour((255,255,255))

       x=13; y=13; 
       sllabel=wx.StaticText(self.blockfunctionswindow,-1,'a factors:',pos=(x,y))
       sllabel.SetFont(self.textfont)
       self.origapanel = wx.Panel(self.blockfunctionswindow, pos=(10, 30), size=(300,300))
       self.panelredrawstatus["origapanel"]=True
       self.origapanel.Bind(wx.EVT_PAINT, lambda event: self.onblockpanelpaint(wx.EVT_PAINT,self.origapanel,'origapanel',"origa"))
  
       x=323; y=13; 
       sllabel=wx.StaticText(self.blockfunctionswindow,-1,'a block functions:',pos=(x,y))
       sllabel.SetFont(self.textfont)
       self.ablockspanel = wx.Panel(self.blockfunctionswindow, pos=(320, 30), size=(300,300))
       self.panelredrawstatus["ablockspanel"]=True
       self.ablockspanel.Bind(wx.EVT_PAINT, lambda event: self.onblockpanelpaint(wx.EVT_PAINT,self.ablockspanel,'ablockspanel',"a"))
  
       x=633; y=13; 
       sllabel=wx.StaticText(self.blockfunctionswindow,-1,'a residuals:',pos=(x,y))
       sllabel.SetFont(self.textfont)
       self.resablockspanel = wx.Panel(self.blockfunctionswindow, pos=(630, 30), size=(300,300))
       self.panelredrawstatus["resablockspanel"]=True
       self.resablockspanel.Bind(wx.EVT_PAINT, lambda event: self.onblockpanelpaint(wx.EVT_PAINT,self.resablockspanel,'resablockspanel',"resa"))
  
       x=13; y=343; 
       sllabel=wx.StaticText(self.blockfunctionswindow,-1,'b factors:',pos=(x,y))
       sllabel.SetFont(self.textfont)
       self.origbpanel = wx.Panel(self.blockfunctionswindow, pos=(10, 360), size=(300,300))
       self.panelredrawstatus["origbpanel"]=True
       self.origbpanel.Bind(wx.EVT_PAINT, lambda event: self.onblockpanelpaint(wx.EVT_PAINT,self.origbpanel,'origbpanel',"origb"))

       x=323; y=343; 
       sllabel=wx.StaticText(self.blockfunctionswindow,-1,'b block functions:',pos=(x,y))
       sllabel.SetFont(self.textfont)
       self.bblockspanel = wx.Panel(self.blockfunctionswindow, pos=(320, 360), size=(300,300))
       self.panelredrawstatus["bblockspanel"]=True
       self.bblockspanel.Bind(wx.EVT_PAINT, lambda event: self.onblockpanelpaint(wx.EVT_PAINT,self.bblockspanel,'bblockspanel',"b"))

       x=633; y=343; 
       sllabel=wx.StaticText(self.blockfunctionswindow,-1,'b residuals:',pos=(x,y))
       sllabel.SetFont(self.textfont)
       self.resbblockspanel = wx.Panel(self.blockfunctionswindow, pos=(630, 360), size=(300,300))
       self.panelredrawstatus["resbblockspanel"]=True
       self.resbblockspanel.Bind(wx.EVT_PAINT, lambda event: self.onblockpanelpaint(wx.EVT_PAINT,self.resbblockspanel,'resbblockspanel',"resb"))


       self.origapanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickAnyBlockPanel)
       self.ablockspanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickAnyBlockPanel)
       self.resablockspanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickAnyBlockPanel)
       self.origbpanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickAnyBlockPanel)
       self.bblockspanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickAnyBlockPanel)
       self.resbblockspanel.Bind(wx.EVT_RIGHT_DOWN, self.OnRightClickAnyBlockPanel)

       self.blockfunctionswindow.Show()

    def onblockpanelpaint(self,evt,panel,panelname,factor,forceredraw=False):

       if self.panelredrawstatus[panelname] or forceredraw:

          # determine first saturday
          firstday=self.minxzoom-self.priorperiodspin.GetValue()
          firstsaturday=5-conversion.xldateweekday(firstday)
          if firstsaturday<0:
             firstsaturday+=7

          # generate data
          nperiod=self.crosstabnperiod

          if factor[-1]=='a':
             table=self.slopedat 
             zerov=1.0
          elif factor[-1]=='b':
             table=self.interceptdat 
             zerov=0.0
          else:
             print('factor error')
             sys.exit(1)

          # compute value range of original matrix
          mint=numpy.amin(table)
          maxt=numpy.amax(table)
          bounds=[mint,maxt]

          if factor[:4]=='orig':
             suppressscalebar=False
          else:
             suppressscalebar=True

          if factor=='a' or factor=='resa':
             table=numpy.zeros([nperiod,nperiod])
             if len(self.clusterstats1)>0:
                templates=blockanalysis.plateautemplates(self.clusterstats1[1],nperiod)
                weekendtemplates=blockanalysis.weekendtemplates(nperiod,firstsaturday)
                table=-1.0*blockanalysis.updateblockdata3o(table,templates,weekendtemplates,self.clusterstats1[0],firstsaturday,self.weekendfactora)

                #table=-1.0*blockanalysis.updateblockdata2o(table,self.clusterstats1[1], self.clusterstats1[0],firstsaturday,self.weekendfactora)
 
             # log -> regular values
             for i in range(len(table)):
               for j in range(len(table)):
                   table[i][j]=10.0**table[i][j]

             # compute residuals
             if factor=='resa':
                table=numpy.subtract(self.slopedat,table)
                bounds=[mint-1.0, maxt-1.0] # residuals are zero-centered
                zerov=0.0
                suppressscalebar=False

          elif factor in ['b', 'resb'] :
             table=numpy.zeros([nperiod,nperiod])
             if len(self.clusterstats2)>0:
                templates=blockanalysis.plateautemplates(self.clusterstats2[1],nperiod)
                weekendtemplates=blockanalysis.weekendtemplates(nperiod,firstsaturday)
                table=-1.0*blockanalysis.updateblockdata3o(table,templates,weekendtemplates,self.clusterstats2[0],firstsaturday,self.weekendfactorb)

                #table=-1.0*blockanalysis.updateblockdata2o(table,self.clusterstats2[1], self.clusterstats2[0],firstsaturday,self.weekendfactorb)
             # compute residuals
             if factor=='resb':
                table=numpy.subtract(self.interceptdat,table)

          if self.monperiodrb.GetValue():
             periodleg='monthly'
          else:
             periodleg='datedays'

          nprior= self.priorperiodspin.GetValue()
          bw=self.ctwblackwhite.GetValue()
          istart=math.floor(self.minxzoom)
          self.crosstabnday=self.periodspin.GetValue()

          if self.setperiodrb.GetValue() and self.crosstabnday==1:
             singledays=True
          else:
             singledays=False

          #print factor+' bounds='+str(bounds)
          [eminv,maxv,self.panelbuffers[panelname]]=visual.drawMatrix(panel,table,self.fitusedat,0,0,300,300,region='U',zerov=zerov,markers=[],bw=bw, bounds=bounds,legend=periodleg,display=True,printvaluerange=True,nprior=nprior,start=istart,singledays=singledays,markercols=[],markerbox=[],label='', suppressscalebar=suppressscalebar)

       else:
          visual.drawBitmap(panel,self.panelbuffers[panelname])
    

    def OnRightClickAnyBlockPanel(self,evt):
       x, y = evt.GetPosition()

       anyblockpanelmenu=wx.Menu()
       ID_SAVEIM=101
       anyblockpanelmenu.Append(ID_SAVEIM, "Save images", "Save images")
       self.Bind(wx.EVT_MENU, lambda event: self.anyblockpanelSaveImage(wx.EVT_MENU,self.blockfunctionswindow), id=ID_SAVEIM)

       self.blockfunctionswindow.PopupMenu( anyblockpanelmenu, evt.GetPosition() )

       anyblockpanelmenu.Destroy() # to prevent memory leak


    def anyblockpanelSaveImage(self,evt,window):
       dlg = wx.FileDialog(self,message='Save image', defaultDir=self.cwd, wildcard='*.png', style=wx.FD_SAVE)
       if dlg.ShowModal() == wx.ID_OK:
          path=dlg.GetPath().replace('.png','').replace('.PNG','')+'.png'
          if not (path==''):
             context = wx.ClientDC( window )
             memory = wx.MemoryDC( )
             x, y = window.ClientSize
             bitmap = wx.EmptyBitmap( x, y, -1 )
             memory.SelectObject( bitmap )
             memory.Blit( 0, 0, x, y, context, 0, 0)
             memory.SelectObject( wx.NullBitmap)
             bitmap.SaveFile( path, wx.BITMAP_TYPE_PNG )
       # bring window back to top of hierarchy
       self.crosstabwindow.Raise()

    def infowin(self,evt):
        lstr=self.appname+'\n'
        lstr+='Copyright (2014, 2023) KWR Watercycle Research Institute\n'
        lstr+='Distribution is of this software is allowed under the\n'
        lstr+='conditions of the UEPL version 1.2\n'
        lstr+='(https://joinup.ec.europa.eu/collection/eupl/eupl-text-eupl-12).\n\n'
        lstr+='Uses the following external libraries:\n'
        lstr+='numpy (www.numpy.org, BSD license)\n' 
        lstr+='xlrd/xlwt (www.python-excel.org, custom license)\n'
        lstr+='wx (www.wxpython.org, LGPL-like license)\n'
        lstr+='pygame (www.pygame.org, LGPL license)\n'
        lstr+='openpyxl (https://foss.heptapod.net/openpyxl/openpyxl, MIT/Expat license)\n\n'
        lstr+='Detailed license info for the tool and these libraries\n'
        lstr+='is included with the program files.\n'
        lstr+='\n'

        dlg = wx.MessageDialog(self, lstr, 'Info', wx.ICON_INFORMATION)
        dlg.ShowModal()

application = wx.App()
window = MyFrame()
application.MainLoop()

